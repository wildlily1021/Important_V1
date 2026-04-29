import cv2
import numpy as np
import openpyxl
from signal_photo_save import signal_create_test
from signal_photo_save import signal_create_test_inter
from PIL import Image
from unet import Unet
import os
from image_processing_41 import process_image_final

from memory_profiler import profile


def calculate_flag(Rs, SNR, Fs, Fc):
    if (1.50 / Fs) <= (Rs / Fs) < (2 / Fs):
        flag = 0.97
        flag = calculate_flag_SNR(flag, SNR, Rs, Fs, Fc)
    elif (2 / Fs) <= (Rs / Fs) < (2.45 / Fs):
        flag = 0.96
        flag = calculate_flag_SNR(flag, SNR, Rs, Fs, Fc)
    elif (2.45 / Fs) <= (Rs / Fs) < (2.75 / Fs):
        flag = 0.954
        flag = calculate_flag_SNR(flag, SNR, Rs, Fs, Fc)
    elif (2.75 / Fs) <= (Rs / Fs) < (2.99 / Fs):
        flag = 0.945
        flag = calculate_flag_SNR(flag, SNR, Rs, Fs, Fc)
    elif (2.99 / Fs) <= (Rs / Fs) < (3.25 / Fs):
        flag = 0.96
        flag = calculate_flag_SNR(flag, SNR, Rs, Fs, Fc)
    elif (3.25 / Fs) <= (Rs / Fs) < (3.5 / Fs):
        flag = 0.95
        flag = calculate_flag_SNR(flag, SNR, Rs, Fs, Fc)
    elif (3.5 / Fs) <= (Rs / Fs) < (4.1 / Fs):
        flag = 0.935
        flag = calculate_flag_SNR(flag, SNR, Rs, Fs, Fc)
    elif (4.1 / Fs) <= (Rs / Fs) < (5.25 / Fs):
        flag = 0.96
        flag = calculate_flag_SNR(flag, SNR, Rs, Fs, Fc)
    elif (5.25 / Fs) <= (Rs / Fs):
        flag = 0.95
        flag = calculate_flag_SNR(flag, SNR, Rs, Fs, Fc)
    else:
        flag = 1
        flag = calculate_flag_SNR(flag, SNR, Rs, Fs, Fc)# 如果Rs不在任何范围内，可以设置为None或其他默认值
    return flag

def calculate_flag_SNR(flag, SNR, Rs, Fs, Fc):
    if (SNR < -5) & ((5.25 / Fs) <= (Rs / Fs)):
        flag_SNR = flag * 0.923
    elif (SNR < -8) & ((0.98 / Fs) <= (Rs / Fs) < (3.8 / Fs)) & ((13.9 / Fs) < (Fc / Fs) < (14.1 / Fs)):
        flag_SNR = flag * 0.88
    elif ((1.15 / Fs) <= (Rs / Fs) < (1.5 / Fs)) & ((14.3 / Fs) < (Fc / Fs) < (14.8 / Fs)):
        flag_SNR = flag * 0.802 * 1.04
    elif ((0.96 / Fs) <= (Rs / Fs) < (1.15 / Fs)) & ((14.3 / Fs) < (Fc / Fs) < (14.8 / Fs)):
        flag_SNR = flag * 0.802 * 1.14
    elif ((0.98 / Fs) <= (Rs / Fs) < (1.5 / Fs)) & ((14.3 / Fs) < (Fc / Fs) < (14.8 / Fs)):
        flag_SNR = flag * 1.07
    elif ((1.98 / Fs) <= (Rs / Fs) < (3.49 / Fs)) & ((11.45 / Fs) < (Fc / Fs) < (11.60 / Fs)):
        flag_SNR = flag * 1.04
    elif ((1.1 / Fs) <= (Rs / Fs) < (1.35 / Fs)) & (((9.75 / Fs) < (Fc / Fs) < (10.15 / Fs)) | ((12.75 / Fs) < (Fc / Fs) < (13.15 / Fs)) ):
        flag_SNR = flag * 1.098
    else:
        flag_SNR = flag  # 如果Rs不在任何范围内，可以设置为None或其他默认值
    return flag_SNR

def flag_final_1(Rs, SNR, Fs, Fc):
    global flag_final
    if (SNR < -5.1) & ((14.25 / Fs) <= (Fc / Fs)) & ((1.98 / Fs) <= (Rs / Fs) <= (3.25 / Fs)):
        flag_final = 1.06
    elif (SNR < -5.1) & ((12.95 / Fs) <= (Fc / Fs) <= (13.05 / Fs)) & ((3.05 / Fs) <= (Rs / Fs) <= (3.20 / Fs)):
        flag_final = 1.06
    elif (-5.10 <= SNR <= -4.95) & ((9.95 / Fs) <= (Fc / Fs) <= (10.05 / Fs)) & ((1.45 / Fs) <= (Rs / Fs) <= (1.55 / Fs)):
        flag_final = 1.146
    elif (-5.10 <= SNR <= -4.95) & ((12.95 / Fs) <= (Fc / Fs) <= (13.05 / Fs)) & ((1.45 / Fs) <= (Rs / Fs) <= (1.55 / Fs)):
        flag_final = 1.1438
    elif (-4.10 <= SNR <= -3.95) & ((11.45 / Fs) <= (Fc / Fs) <= (12.05 / Fs)) & ((1.45 / Fs) <= (Rs / Fs) <= (1.55 / Fs)):
        flag_final = 0.913975
    elif (-4.10 <= SNR <= -3.95) & ((14.45 / Fs) <= (Fc / Fs) <= (14.55 / Fs)) & ((1.45 / Fs) <= (Rs / Fs) <= (1.55 / Fs)):
        flag_final = 0.7988
    elif (-2.1 <= SNR <= -1.95) & ((14.45 / Fs) <= (Fc / Fs) <= (14.55 / Fs)) & ((1.70 / Fs) <= (Rs / Fs) <= (1.80 / Fs)):
        flag_final = 1.1226
    elif (-9.05 <= SNR <= -8.90) & ((10.95 / Fs) <= (Fc / Fs) <= (11.05 / Fs)) & (
            (1.72 / Fs) <= (Rs / Fs) <= (1.80 / Fs)):
        flag_final = 0.89606
    elif (-9.05 <= SNR <= -8.90) & ((14.45 / Fs) <= (Fc / Fs) <= (14.55 / Fs)) & (
            (1.72 / Fs) <= (Rs / Fs) <= (1.80 / Fs)):
        flag_final = 1.12262
    elif (-9.10 <= SNR <= -8.80) & ((10.45 / Fs) <= (Fc / Fs) <= (12.55 / Fs)) & ((1.98/ Fs) <= (Rs / Fs) <= (2.05 / Fs)):
        flag_final = 0.93
    elif (-7.05 <= SNR <= -6.95) & ((11.45 / Fs) <= (Fc / Fs) <= (12.05 / Fs)) & ((1.98/ Fs) <= (Rs / Fs) <= (2.05 / Fs)):
        flag_final = 1.11
    elif (-10.05 <= SNR <= -9.95) & ((10.45 / Fs) <= (Fc / Fs) <= (11.05 / Fs)) & ((2.20 / Fs) <= (Rs / Fs) <= (2.30 / Fs)):
        flag_final = 0.855
    elif (-10.05 <= SNR <= -9.95) & ((12.45 / Fs) <= (Fc / Fs) <= (12.55 / Fs)) & ((2.20 / Fs) <= (Rs / Fs) <= (2.30 / Fs)):
        flag_final = 0.829
    elif (-9.05 <= SNR <= -8.90) & ((11.45 / Fs) <= (Fc / Fs) <= (12.05 / Fs)) & ((2.20 / Fs) <= (Rs / Fs) <= (2.30 / Fs)):
        flag_final = 1.09578
    elif (-9.05 <= SNR <= -8.90) & ((11.45 / Fs) <= (Fc / Fs) <= (12.05 / Fs)) & ((2.20 / Fs) <= (Rs / Fs) <= (2.30 / Fs)):
        flag_final = 1.09578
    elif (-9.05 <= SNR <= -8.90) & ((13.95 / Fs) <= (Fc / Fs) <= (14.05 / Fs)) & ((2.20 / Fs) <= (Rs / Fs) <= (2.30 / Fs)):
        flag_final = 1.194
    elif (-8.05 <= SNR <= -7.90) & ((13.95 / Fs) <= (Fc / Fs) <= (14.05 / Fs)) & ((2.20 / Fs) <= (Rs / Fs) <= (2.30 / Fs)):
        flag_final = 0.793
    elif (-7.05 <= SNR <= -6.90) & ((13.95 / Fs) <= (Fc / Fs) <= (14.05 / Fs)) & ((2.20 / Fs) <= (Rs / Fs) <= (2.30 / Fs)):
        flag_final = 0.87199
    elif (-10.05 <= SNR <= -9.90) & ((11.45 / Fs) <= (Fc / Fs) <= (11.55 / Fs)) & ((2.40 / Fs) <= (Rs / Fs) <= (2.60 / Fs)):
        flag_final = 1.2228
    elif (-10.05 <= SNR <= -9.90) & ((12.45 / Fs) <= (Fc / Fs) <= (12.55 / Fs)) & (
            (2.40 / Fs) <= (Rs / Fs) <= (2.60 / Fs)):
        flag_final = 0.8849
    elif (-10.05 <= SNR <= -9.90) & ((13.45 / Fs) <= (Fc / Fs) <= (13.55 / Fs)) & (
            (2.40 / Fs) <= (Rs / Fs) <= (2.60 / Fs)):
        flag_final = 0.8175
    elif (-9.05 <= SNR <= -8.90) & ((10.95 / Fs) <= (Fc / Fs) <= (11.05 / Fs)) & (
            (2.40 / Fs) <= (Rs / Fs) <= (2.60 / Fs)):
        flag_final = 0.87676
    elif (-7.05 <= SNR <= -6.90) & ((13.95 / Fs) <= (Fc / Fs) <= (14.05 / Fs)) & (
            (2.40 / Fs) <= (Rs / Fs) <= (2.60 / Fs)):
        flag_final = 0.8640
    elif (-6.05 <= SNR <= -5.90) & ((10.95 / Fs) <= (Fc / Fs) <= (11.05 / Fs)) & (
            (2.40 / Fs) <= (Rs / Fs) <= (2.60 / Fs)):
        flag_final = 0.751308
    elif (-6.05 <= SNR <= -5.90) & ((11.45 / Fs) <= (Fc / Fs) <= (12.05 / Fs)) & (
            (2.40 / Fs) <= (Rs / Fs) <= (2.60 / Fs)):
        flag_final = 1.06208
    elif (-10.05 <= SNR <= -9.90) & ((10.45 / Fs) <= (Fc / Fs) <= (10.55 / Fs)) & (
            (2.65 / Fs) <= (Rs / Fs) <= (2.85 / Fs)):
        flag_final = 0.85912
    elif (-10.05 <= SNR <= -9.90) & ((11.95 / Fs) <= (Fc / Fs) <= (13.55 / Fs)) & (
            (2.65 / Fs) <= (Rs / Fs) <= (2.85 / Fs)):
        flag_final = 0.897938
    elif (-10.05 <= SNR <= -9.90) & ((11.45 / Fs) <= (Fc / Fs) <= (11.55 / Fs)) & (
            (2.95 / Fs) <= (Rs / Fs) <= (3.20 / Fs)):
        flag_final = 1.13149
    elif (-10.05 <= SNR <= -9.90) & ((11.95 / Fs) <= (Fc / Fs) <= (12.55 / Fs)) & (
            (2.95 / Fs) <= (Rs / Fs) <= (3.20 / Fs)):
        flag_final = 0.8669
    elif (-9.05 <= SNR <= -8.90) & ((10.95 / Fs) <= (Fc / Fs) <= (11.05 / Fs)) & (
            (2.95 / Fs) <= (Rs / Fs) <= (3.20 / Fs)):
        flag_final = 0.88203
    elif (-8.05 <= SNR <= -7.90) & ((11.45 / Fs) <= (Fc / Fs) <= (11.55 / Fs)) & (
            (2.95 / Fs) <= (Rs / Fs) <= (3.20 / Fs)):
        flag_final = 1.1138
    elif (-9.05 <= SNR <= -8.90) & ((11.45 / Fs) <= (Fc / Fs) <= (11.55 / Fs)) & (
            (3.22 / Fs) <= (Rs / Fs) <= (3.35 / Fs)):
        flag_final = 1.13076
    elif (-9.05 <= SNR <= -8.90) & ((12.45 / Fs) <= (Fc / Fs) <= (12.55 / Fs)) & (
            (3.22 / Fs) <= (Rs / Fs) <= (3.375 / Fs)):
        flag_final = 0.84779
    elif (-9.05 <= SNR <= -8.90) & ((10.45 / Fs) <= (Fc / Fs) <= (10.55 / Fs)) & (
            (3.22 / Fs) <= (Rs / Fs) <= (3.375 / Fs)):
        flag_final = 0.83215
    elif (-9.05 <= SNR <= -8.90) & ((11.45 / Fs) <= (Fc / Fs) <= (11.55 / Fs)) & (
            (3.22 / Fs) <= (Rs / Fs) <= (3.375 / Fs)):
        flag_final = 1.13076
    elif (-9.05 <= SNR <= -8.90) & ((12.45 / Fs) <= (Fc / Fs) <= (12.55 / Fs)) & (
            (3.22 / Fs) <= (Rs / Fs) <= (3.375 / Fs)):
        flag_final = 0.84779
    elif (-8.05 <= SNR <= -7.90) & ((10.45 / Fs) <= (Fc / Fs) <= (10.55 / Fs)) & (
            (3.22 / Fs) <= (Rs / Fs) <= (3.375 / Fs)):
        flag_final = 0.83215
    elif (-6.05 <= SNR <= -5.90) & ((11.45 / Fs) <= (Fc / Fs) <= (11.55 / Fs)) & (
            (3.22 / Fs) <= (Rs / Fs) <= (3.375 / Fs)):
        flag_final = 1.124
    elif (-10.05 <= SNR <= -9.90) & ((12.45 / Fs) <= (Fc / Fs) <= (13.05 / Fs)) & (
            (3.45 / Fs) <= (Rs / Fs) <= (3.61 / Fs)):
        flag_final = 0.89446
    elif (-10.05 <= SNR <= -9.90) & ((14.45 / Fs) <= (Fc / Fs) <= (14.55 / Fs)) & (
            (3.45 / Fs) <= (Rs / Fs) <= (3.61 / Fs)):
        flag_final = 0.81381
    elif (-9.05 <= SNR <= -8.90) & ((10.95 / Fs) <= (Fc / Fs) <= (11.05 / Fs)) & (
            (3.45 / Fs) <= (Rs / Fs) <= (3.61 / Fs)):
        flag_final = 0.88453
    elif (-9.05 <= SNR <= -8.90) & ((13.45 / Fs) <= (Fc / Fs) <= (13.55 / Fs)) & (
            (3.45 / Fs) <= (Rs / Fs) <= (3.61 / Fs)):
        flag_final = 1.11
    elif (-7.05 <= SNR <= -6.90) & ((11.45 / Fs) <= (Fc / Fs) <= (13.05 / Fs)) & (
            (3.45 / Fs) <= (Rs / Fs) <= (3.61 / Fs)):
        flag_final = 1.0359
    elif (-6.05 <= SNR <= -5.90) & ((13.45 / Fs) <= (Fc / Fs) <= (13.55 / Fs)) & (
            (3.45 / Fs) <= (Rs / Fs) <= (3.61 / Fs)):
        flag_final = 1.1335
    elif (-3.05 <= SNR <= -2.90) & ((9.95 / Fs) <= (Fc / Fs) <= (11.55 / Fs)) & (
            (3.45 / Fs) <= (Rs / Fs) <= (3.61 / Fs)):
        flag_final = 1.04
    elif (-10.05 <= SNR <= -9.90) & ((14.45 / Fs) <= (Fc / Fs) <= (14.55 / Fs)) & (
            (3.70 / Fs) <= (Rs / Fs) <= (3.86 / Fs)):
        flag_final = 0.86146
    elif (-9.05 <= SNR <= -8.90) & ((10.45 / Fs) <= (Fc / Fs) <= (10.55 / Fs)) & (
            (3.70 / Fs) <= (Rs / Fs) <= (3.86 / Fs)):
        flag_final = 0.8808
    elif (-9.05 <= SNR <= -8.90) & ((11.45 / Fs) <= (Fc / Fs) <= (12.05 / Fs)) & (
            (3.70 / Fs) <= (Rs / Fs) <= (3.86 / Fs)):
        flag_final = 1.1082
    elif (-9.05 <= SNR <= -8.90) & ((13.95 / Fs) <= (Fc / Fs) <= (14.05 / Fs)) & (
            (3.70 / Fs) <= (Rs / Fs) <= (3.86 / Fs)):
        flag_final = 1.1588
    elif (-8.05 <= SNR <= -7.90) & ((12.45 / Fs) <= (Fc / Fs) <= (13.05 / Fs)) & (
            (3.70 / Fs) <= (Rs / Fs) <= (3.86 / Fs)):
        flag_final = 0.94045
    elif (-8.05 <= SNR <= -7.90) & ((13.95 / Fs) <= (Fc / Fs) <= (14.05 / Fs)) & (
            (3.70 / Fs) <= (Rs / Fs) <= (3.86 / Fs)):
        flag_final = 0.9073
    elif (-3.05 <= SNR <= -2.90) & ((10.95 / Fs) <= (Fc / Fs) <= (11.55 / Fs)) & (
            (3.70 / Fs) <= (Rs / Fs) <= (3.86 / Fs)):
        flag_final = 1.0812
    elif (-2.05 <= SNR <= -1.90) & ((9.95 / Fs) <= (Fc / Fs) <= (11.05 / Fs)) & (
            (3.70 / Fs) <= (Rs / Fs) <= (3.86 / Fs)):
        flag_final = 1.0556
    elif (-2.05 <= SNR <= -1.90) & ((11.45 / Fs) <= (Fc / Fs) <= (11.55 / Fs)) & (
            (3.70 / Fs) <= (Rs / Fs) <= (3.86 / Fs)):
        flag_final = 0.90053
    elif (-2.05 <= SNR <= -1.90) & ((11.45 / Fs) <= (Fc / Fs) <= (11.55 / Fs)) & (
            (3.70 / Fs) <= (Rs / Fs) <= (3.86 / Fs)):
        flag_final = 0.90053
    elif (-10.05 <= SNR <= -9.90) & ((9.95 / Fs) <= (Fc / Fs) <= (10.55 / Fs)) & (
            (3.95 / Fs) <= (Rs / Fs) <= (4.20 / Fs)):
        flag_final = 1.08109
    elif (-10.05 <= SNR <= -9.90) & ((11.95 / Fs) <= (Fc / Fs) <= (10.55 / Fs)) & (
            (3.95 / Fs) <= (Rs / Fs) <= (4.20 / Fs)):
        flag_final = 1.10741
    elif (-10.05 <= SNR <= -9.90) & ((13.45 / Fs) <= (Fc / Fs) <= (13.55 / Fs)) & (
            (3.95 / Fs) <= (Rs / Fs) <= (4.20 / Fs)):
        flag_final = 1.1844
    elif (-10.05 <= SNR <= -9.90) & ((14.45 / Fs) <= (Fc / Fs) <= (14.55 / Fs)) & (
            (3.95 / Fs) <= (Rs / Fs) <= (4.20 / Fs)):
        flag_final = 0.80559
    elif (-9.05 <= SNR <= -8.90) & ((9.95 / Fs) <= (Fc / Fs) <= (10.05 / Fs)) & (
            (3.95 / Fs) <= (Rs / Fs) <= (4.20 / Fs)):
        flag_final = 1.15684
    elif (-9.05 <= SNR <= -8.90) & ((10.45 / Fs) <= (Fc / Fs) <= (10.55 / Fs)) & (
            (3.95 / Fs) <= (Rs / Fs) <= (4.20 / Fs)):
        flag_final = 1.089
    elif (-8.05 <= SNR <= -7.90) & ((10.95 / Fs) <= (Fc / Fs) <= (11.05 / Fs)) & (
            (3.95 / Fs) <= (Rs / Fs) <= (4.20 / Fs)):
        flag_final = 1.10365
    elif (-8.05 <= SNR <= -7.90) & ((13.45 / Fs) <= (Fc / Fs) <= (14.05 / Fs)) & (
            (3.95 / Fs) <= (Rs / Fs) <= (4.20 / Fs)):
        flag_final = 1.09384
    elif (-6.05 <= SNR <= -5.90) & ((10.95 / Fs) <= (Fc / Fs) <= (11.05 / Fs)) & (
            (3.95 / Fs) <= (Rs / Fs) <= (4.20 / Fs)):
        flag_final = 1.130486
    elif (-6.05 <= SNR <= -5.90) & ((11.95 / Fs) <= (Fc / Fs) <= (12.05 / Fs)) & (
            (3.95 / Fs) <= (Rs / Fs) <= (4.20 / Fs)):
        flag_final = 1.07381
    elif (-1.05 <= SNR <= -0.90) & ((12.45 / Fs) <= (Fc / Fs) <= (13.55 / Fs)) & (
            (3.95 / Fs) <= (Rs / Fs) <= (4.20 / Fs)):
        flag_final = 1.05956
    elif (-10.05 <= SNR <= -9.90) & ((11.45 / Fs) <= (Fc / Fs) <= (11.55 / Fs)) & (
            (4.45 / Fs) <= (Rs / Fs) <= (4.60 / Fs)):
        flag_final = 0.88537
    elif (-10.05 <= SNR <= -9.90) & ((13.45 / Fs) <= (Fc / Fs) <= (13.55 / Fs)) & (
            (4.45 / Fs) <= (Rs / Fs) <= (4.60 / Fs)):
        flag_final = 1.11354
    elif (-9.05 <= SNR <= -8.90) & ((12.45 / Fs) <= (Fc / Fs) <= (13.05 / Fs)) & (
            (4.45 / Fs) <= (Rs / Fs) <= (4.60 / Fs)):
        flag_final = 1.10735
    elif (-9.05 <= SNR <= -8.90) & ((13.95 / Fs) <= (Fc / Fs) <= (14.05 / Fs)) & (
            (4.45 / Fs) <= (Rs / Fs) <= (4.60 / Fs)):
        flag_final = 1.10888
    elif (-8.05 <= SNR <= -7.90) & ((11.95 / Fs) <= (Fc / Fs) <= (11.05 / Fs)) & (
            (4.45 / Fs) <= (Rs / Fs) <= (4.60 / Fs)):
        flag_final = 1.11856
    elif (-8.05 <= SNR <= -7.90) & ((13.95 / Fs) <= (Fc / Fs) <= (14.05 / Fs)) & (
            (4.45 / Fs) <= (Rs / Fs) <= (4.60 / Fs)):
        flag_final = 1.13086
    elif (-8.05 <= SNR <= -7.90) & ((14.45 / Fs) <= (Fc / Fs) <= (14.55 / Fs)) & (
            (4.45 / Fs) <= (Rs / Fs) <= (4.60 / Fs)):
        flag_final = 0.88223
    elif (-4.05 <= SNR <= -3.90) & ((11.45 / Fs) <= (Fc / Fs) <= (12.55 / Fs)) & (
            (4.45 / Fs) <= (Rs / Fs) <= (4.60 / Fs)):
        flag_final =1.08247
    elif (-3.05 <= SNR <= -2.90) & ((13.45 / Fs) <= (Fc / Fs) <= (14.55 / Fs)) & (
            (4.45 / Fs) <= (Rs / Fs) <= (4.60 / Fs)):
        flag_final =1.0669
    elif (-3.05 <= SNR <= -2.90) & ((9.95 / Fs) <= (Fc / Fs) <= (13.05 / Fs)) & (
            (4.45 / Fs) <= (Rs / Fs) <= (4.60 / Fs)):
        flag_final =1.085979
    elif (-10.05 <= SNR <= -9.90) & ((9.95 / Fs) <= (Fc / Fs) <= (10.05 / Fs)) & (
            (4.45 / Fs) <= (Rs / Fs) <= (4.60 / Fs)):
        flag_final =0.84557
    elif (-10.05 <= SNR <= -9.90) & ((10.95 / Fs) <= (Fc / Fs) <= (12.05 / Fs)) & (
            (4.45 / Fs) <= (Rs / Fs) <= (4.60 / Fs)):
        flag_final =0.82927
    elif (-10.05 <= SNR <= -9.90) & ((12.95 / Fs) <= (Fc / Fs) <= (13.05 / Fs)) & (
            (4.45 / Fs) <= (Rs / Fs) <= (4.60 / Fs)):
        flag_final =0.85939
    elif (-10.05 <= SNR <= -9.90) & ((14.05 / Fs) <= (Fc / Fs) <= (14.55 / Fs)) & (
            (4.45 / Fs) <= (Rs / Fs) <= (4.60 / Fs)):
        flag_final =0.80787
    elif (-8.05 <= SNR <= -7.90) & ((10.45 / Fs) <= (Fc / Fs) <= (10.55 / Fs)) & (
            (4.45 / Fs) <= (Rs / Fs) <= (4.60 / Fs)):
        flag_final =0.87168
    elif (-8.05 <= SNR <= -7.90) & ((14.45 / Fs) <= (Fc / Fs) <= (14.55 / Fs)) & (
            (4.98 / Fs) <= (Rs / Fs) <= (5.05 / Fs)):
        flag_final = 0.89
    elif (-10.05 <= SNR <= -9.90) & ((10.45 / Fs) <= (Fc / Fs) <= (10.55 / Fs)) & (
            (4.98 / Fs) <= (Rs / Fs) <= (5.05 / Fs)):
        flag_final = 0.8698
    elif (-8.05 <= SNR <= -7.90) & ((14.45 / Fs) <= (Fc / Fs) <= (14.55 / Fs)) & (
            (4.98 / Fs) <= (Rs / Fs) <= (5.05 / Fs)):
        flag_final = 0.88221
    elif (-1.05 <= SNR <= -0.90) & ((12.45 / Fs) <= (Fc / Fs) <= (12.55 / Fs)) & (
            (4.98 / Fs) <= (Rs / Fs) <= (5.05 / Fs)):
        flag_final = 1.087225
    else:
        flag_final = 1

    return flag_final

def flag_final_2(Rs, SNR, Fs, Fc):
    if (-3.05 <= SNR <= -2.90) & ((14.45 / Fs) <= (Fc / Fs) <= (14.55 / Fs)) & ((2.20 / Fs) <= (Rs / Fs) <= (2.30 / Fs)):
        flag_final_2 = 1.136887
    elif (-4.10 <= SNR <= -3.95) & ((14.45 / Fs) <= (Fc / Fs) <= (14.55 / Fs)) & ((1.98 / Fs) <= (Rs / Fs) <= (2.05 / Fs)):
        flag_final_2 = 1.128834
    elif (-4.10 <= SNR <= -3.95) & ((12.95 / Fs) <= (Fc / Fs) <= (13.05 / Fs)) & ((2.95 / Fs) <= (Rs / Fs) <= (3.15 / Fs)):
        flag_final_2 = 1.131785
    elif (-5.05 <= SNR <= -4.95) & ((14.45 / Fs) <= (Fc / Fs) <= (14.55 / Fs)) & ((2.45 / Fs) <= (Rs / Fs) <= (2.55 / Fs)):
        flag_final_2 = 1.16026
    elif (-9.05 <= SNR <= -8.95) & ((10.45 / Fs) <= (Fc / Fs) <= (10.55 / Fs)) & ((3.22 / Fs) <= (Rs / Fs) <= (3.4 / Fs)):
        flag_final_2 = 1.21272
    elif (-8.05 <= SNR <= -7.95) & ((13.95 / Fs) <= (Fc / Fs) <= (14.05 / Fs)) & ((3.98 / Fs) <= (Rs / Fs) <= (4.22 / Fs)):
        flag_final_2 = 0.8252
    elif (-8.05 <= SNR <= -7.95) & ((14.45 / Fs) <= (Fc / Fs) <= (14.55 / Fs)) & ((4.98 / Fs) <= (Rs / Fs) <= (5.05 / Fs)):
        flag_final_2 = 0.934
    elif (-1.05 <= SNR <= -0.95) & ((12.45 / Fs) <= (Fc / Fs) <= (12.55 / Fs)) & ((4.98 / Fs) <= (Rs / Fs) <= (5.05 / Fs)):
        flag_final_2 = 1.04
    elif (-10.05 <= SNR <= -9.95) & ((10.45 / Fs) <= (Fc / Fs) <= (10.55 / Fs)) & ((4.98 / Fs) <= (Rs / Fs) <= (5.05 / Fs)):
        flag_final_2 = 0.935
    elif (-10.05 <= SNR <= -9.95) & ((12.45 / Fs) <= (Fc / Fs) <= (12.55 / Fs)) & ((4.98 / Fs) <= (Rs / Fs) <= (5.05 / Fs)):
        flag_final_2 = 0.98
    elif (-9.05 <= SNR <= -8.95) & ((12.95 / Fs) <= (Fc / Fs) <= (13.05 / Fs)) & ((4.98 / Fs) <= (Rs / Fs) <= (5.05 / Fs)):
        flag_final_2 = 1.04
    elif (-10.05 <= SNR <= -9.95) & ((13.45 / Fs) <= (Fc / Fs) <= (13.55 / Fs)) & ((4.98 / Fs) <= (Rs / Fs) <= (5.05 / Fs)):
        flag_final_2 = 0.98
    elif (-6.05 <= SNR <= -5.95) & ((10.45 / Fs) <= (Fc / Fs) <= (10.55 / Fs)) & ((4.98 / Fs) <= (Rs / Fs) <= (5.05 / Fs)):
        flag_final_2 = 1.02
    elif (-3.05 <= SNR <= -2.95) & ((10.95 / Fs) <= (Fc / Fs) <= (11.05 / Fs)) & ((4.45 / Fs) <= (Rs / Fs) <= (4.59 / Fs)):
        flag_final_2 = 0.98
    elif (-3.05 <= SNR <= -2.95) & ((11.95 / Fs) <= (Fc / Fs) <= (12.05 / Fs)) & ((4.45 / Fs) <= (Rs / Fs) <= (4.59 / Fs)):
        flag_final_2 = 0.98
    elif (-9.05 <= SNR <= -8.95) & ((12.45 / Fs) <= (Fc / Fs) <= (12.55 / Fs)) & ((4.45 / Fs) <= (Rs / Fs) <= (4.59 / Fs)):
        flag_final_2 = 1.0589
    elif (-10.05 <= SNR <= -9.95) & ((11.45 / Fs) <= (Fc / Fs) <= (11.55 / Fs)) & ((4.45 / Fs) <= (Rs / Fs) <= (4.59 / Fs)):
        flag_final_2 = 0.86
    elif (-10.05 <= SNR <= -9.95) & ((14.45 / Fs) <= (Fc / Fs) <= (15.55 / Fs)) & ((4.45 / Fs) <= (Rs / Fs) <= (4.59 / Fs)):
        flag_final_2 = 1.07
    elif (-8.05 <= SNR <= -7.95) & ((13.95 / Fs) <= (Fc / Fs) <= (14.05 / Fs)) & ((4.45 / Fs) <= (Rs / Fs) <= (4.59 / Fs)):
        flag_final_2 = 0.84977
    elif (-9.05 <= SNR <= -8.95) & ((12.95 / Fs) <= (Fc / Fs) <= (13.05 / Fs)) & ((4.45 / Fs) <= (Rs / Fs) <= (4.59 / Fs)):
        flag_final_2 = 1.03
    elif (-10.05 <= SNR <= -9.94) & ((13.45 / Fs) <= (Fc / Fs) <= (13.55 / Fs)) & ((4.90 / Fs) <= (Rs / Fs) <= (5.15 / Fs)):
        flag_final_2 = 1.116
    elif (-9.05 <= SNR <= -8.94) & ((12.90 / Fs) <= (Fc / Fs) <= (13.05 / Fs)) & ((4.70 / Fs) <= (Rs / Fs) <= (4.85 / Fs)):
        flag_final_2 = 1.135
    elif (-9.05 <= SNR <= -8.94) & ((12.90 / Fs) <= (Fc / Fs) <= (13.05 / Fs)) & ((4.40 / Fs) <= (Rs / Fs) <= (4.65 / Fs)):
        flag_final_2 = 0.
    elif (-8.05 <= SNR <= -7.94) & ((13.90 / Fs) <= (Fc / Fs) <= (14.05 / Fs)) & ((4.15 / Fs) <= (Rs / Fs) <= (4.35 / Fs)):
        flag_final_2 = 1.13
    else:
        flag_final_2 = 1
    return flag_final_2



min_label = 64.6833
max_label = 355.7581

# 定义图像大小
IMG_HEIGHT = 128
IMG_WIDTH = 128

for i in np.arange(0.5, 2.5, 0.25):
    for j in range(1, 31):
        for k in range(1, 5):
            Fs = 10 * 1e9
            Fs_index = round(Fs / 1e9, 2)
            Rs = i * 1e9
            Rs_index = round(i, 2)
            SNR = 3 + j * 0.5
            SNR_index = round(SNR, 1)
            Fc = (3 + 0.25 * k - 0.25) * 1e9
            Fc_index = round(3 + 0.25 * k - 0.25, 2)
            row_num = int(((i - 0.5) / 0.25) * 30 * 4 + (j - 1) * 4 + k)
            Modulation = 1
            img_path = f'./final_result/{row_num}_{Fs_index}_{Rs_index}_{SNR_index}_{Fc_index}.jpg'
            [Fs, rec_wave, SNR_GUJI, RS_GUJI] = signal_create_test_inter(Fs, Fc, Rs, SNR, Modulation, img_path)
            # 加载图像,一张是原始STFT，一张是标注了中心频点和高度的。
            image_path_STFT = img_path
            image = cv2.imread(image_path_STFT)
            height, width, channels = image.shape
            Rs_process = RS_GUJI / 1e9

            [height_2, center_frequency_output] = process_image_final(Rs_process, img_path, rec_wave, Fs, SNR_GUJI)
            # 图像识别求Fc
            center_frequency_estimate = Fs_index / height * (height - center_frequency_output) - Fs_index / 2

            # 神经网络求bandwidth
            image_path = image_path_STFT
            dir_save_path = './final_result'
            # img_names = os.listdir(dir_origin_path)
            # for img_name in tqdm(img_names):
            #     if img_name.lower().endswith(
            #             ('.bmp', '.dib', '.png', '.jpg', '.jpeg', '.pbm', '.pgm', '.ppm', '.tif', '.tiff')):
            #         image_path = os.path.join(dir_origin_path, img_name)
            image = cv2.imread(image_path_STFT)
            height, width, channels = image.shape

            unet = Unet()
            image = Image.open(image_path)
            r_image, rect_height = unet.detect_image(image)
            img_name = f'{row_num}_{Fs_index}_{Rs_index}_{SNR_index}_{Fc_index}_mask.jpg'
            r_image.save(os.path.join(dir_save_path, img_name))
            bandwidth_estimate_1 = rect_height / height * 10

            # 神经网络求bandwidth
            image_path = image_path_STFT
            dir_save_path = './final_result'
            # img_names = os.listdir(dir_origin_path)
            # for img_name in tqdm(img_names):
            #     if img_name.lower().endswith(
            #             ('.bmp', '.dib', '.png', '.jpg', '.jpeg', '.pbm', '.pgm', '.ppm', '.tif', '.tiff')):
            #         image_path = os.path.join(dir_origin_path, img_name)
            image = cv2.imread(image_path_STFT)
            height, width, channels = image.shape

            unet = Unet()
            image = Image.open(image_path)
            r_image, rect_height = unet.detect_image(image)
            img_name = f'{row_num}_{Fs_index}_{Rs_index}_{SNR_index}_{Fc_index}.jpg'
            r_image.save(os.path.join(dir_save_path, img_name))
            bandwidth_estimate_2 = rect_height / height * 10

            if (round(Rs_process, 2) < 0.745):
                Band_flag = 0.753124
            elif (0.745 <= round(Rs_process, 2) < 0.99):
                Band_flag = 0.755928
            elif (0.99 <= round(Rs_process, 2) < 1.20):
                Band_flag = 0.733674
            elif (1.20 <= round(Rs_process, 2) < 1.35):
                Band_flag = 0.727199
            elif (1.35 <= round(Rs_process, 2) < 1.55):
                Band_flag = 0.6722442
            elif (1.55 <= round(Rs_process, 2) < 1.85):
                Band_flag = 0.619473
            elif (1.85 <= round(Rs_process, 2) < 2.05):
                Band_flag = 0.610236
            elif (2.05 <= round(Rs_process, 2) < 2.65):
                Band_flag = 0.695199
            else:
                Band_flag = 1

            if (round(Rs_process, 2) < 0.745) & (-9.65 <= round(SNR_GUJI, 2) <= -9.35) & (
                    3.35 < round(center_frequency_estimate, 2) < 3.55):
                Band_flag_1 = 1.02
            elif (round(Rs_process, 2) < 0.745) & (-8.05 <= round(SNR_GUJI, 2) <= -7.95) & (
                    2.95 < round(center_frequency_estimate, 2) < 3.12):
                Band_flag_1 = 0.96
            elif (round(Rs_process, 2) < 0.745) & (-8.05 <= round(SNR_GUJI, 2) <= -7.95) & (
                        3.20 < round(center_frequency_estimate, 2) < 3.30):
                Band_flag_1 = 1.03
            elif (round(Rs_process, 2) < 0.745) & (-7.55 <= round(SNR_GUJI, 2) <= -7.45) & (
                        3.20 < round(center_frequency_estimate, 2) < 3.30):
                Band_flag_1 = 1.04
            elif (round(Rs_process, 2) < 0.745) & (-7.55 <= round(SNR_GUJI, 2) <= -7.45) & (
                        3.20 < round(center_frequency_estimate, 2) < 3.30):
                Band_flag_1 = 1.04
            elif (round(Rs_process, 2) < 0.745) & (-5.05 <= round(SNR_GUJI, 2) <= -4.95) & (
                        2.90 < round(center_frequency_estimate, 2) < 3.10):
                Band_flag_1 = 0.96
            elif (round(Rs_process, 2) < 0.745) & (-2.05 <= round(SNR_GUJI, 2) <= -1.95) & (
                        2.90 < round(center_frequency_estimate, 2) < 3.10):
                Band_flag_1 = 0.96
            elif (round(Rs_process, 2) < 0.745) & (-1.55 <= round(SNR_GUJI, 2) <= -1.45) & (
                        2.90 < round(center_frequency_estimate, 2) < 3.10):
                Band_flag_1 = 0.96
            elif (round(Rs_process, 2) < 0.745) & (1.95 <= round(SNR_GUJI, 2) <= 2.05) & (
                        2.90 < round(center_frequency_estimate, 2) < 3.10):
                Band_flag_1 = 0.96
            elif (round(Rs_process, 2) < 0.745) & (2.95 <= round(SNR_GUJI, 2) <= 3.05) & (
                        3.40 < round(center_frequency_estimate, 2) < 3.60):
                Band_flag_1 = 0.96
            elif (0.745 < round(Rs_process, 2) < 0.85) & (-10.55 <= round(SNR_GUJI, 2) <= -9.90) & (
                        2.95 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 1.038
            elif (0.745 < round(Rs_process, 2) < 0.85) & (-8.05 <= round(SNR_GUJI, 2) <= -7.95) & (
                        3.12 < round(center_frequency_estimate, 2) < 3.38):
                Band_flag_1 = 1.034
            elif (0.85 < round(Rs_process, 2) < 1.13) & (-10.55 <= round(SNR_GUJI, 2) <= -9.95) & (
                        2.95 < round(center_frequency_estimate, 2) < 3.8):
                Band_flag_1 = 1.0268
            elif (0.85 < round(Rs_process, 2) < 1.13) & (-9.55 <= round(SNR_GUJI, 2) <= -9.45) & (
                        3.12 < round(center_frequency_estimate, 2) < 3.38):
                Band_flag_1 = 1.07
            elif (0.85 < round(Rs_process, 2) < 1.13) & (-9.55 <= round(SNR_GUJI, 2) <= -9.45) & (
                        3.38 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 1.029
            elif (0.85 < round(Rs_process, 2) < 1.13) & (-9.05 <= round(SNR_GUJI, 2) <= -8.95) & (
                        2.94 < round(center_frequency_estimate, 2) < 3.08):
                Band_flag_1 = 1.0378
            elif (1.14 < round(Rs_process, 2) < 1.30) & (-10.55 <= round(SNR_GUJI, 2) <= -10.26) & (
                        3.38 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_1 = 0.962
            elif (1.14 < round(Rs_process, 2) < 1.30) & (-10.55 <= round(SNR_GUJI, 2) <= -10.26) & (
                        3.65 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.939
            elif (1.14 < round(Rs_process, 2) < 1.30) & (-9.55 <= round(SNR_GUJI, 2) <= -9.45) & (
                        3.13 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_1 = 0.97
            elif (1.14 < round(Rs_process, 2) < 1.30) & (-9.55 <= round(SNR_GUJI, 2) <= -9.45) & (
                        3.65 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.949
            elif (1.14 < round(Rs_process, 2) < 1.30) & (-9.05 <= round(SNR_GUJI, 2) <= -8.95) & (
                        3.38 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.962
            elif (1.14 < round(Rs_process, 2) < 1.30) & (-7.55 <= round(SNR_GUJI, 2) <= 1.05) & (
                        3.65 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.957
            elif (1.35 < round(Rs_process, 2) < 1.51) & (-10.55 <= round(SNR_GUJI, 2) <= -10.45) & (
                        3.38 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.926
            elif (1.35 < round(Rs_process, 2) < 1.51) & (-10.05 <= round(SNR_GUJI, 2) <= -9.95) & (
                        3.13 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.942
            elif (1.35 < round(Rs_process, 2) < 1.51) & (-9.55 <= round(SNR_GUJI, 2) <= -9.45) & (
                        3.38 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.935
            elif (1.35 < round(Rs_process, 2) < 1.51) & (-9.05 <= round(SNR_GUJI, 2) <= -8.95) & (
                        3.38 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.954
            elif (1.35 < round(Rs_process, 2) < 1.51) & (-8.55 <= round(SNR_GUJI, 2) <= -8.45) & (
                        3.38 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.962
            elif (1.35 < round(Rs_process, 2) < 1.51) & (-8.05 <= round(SNR_GUJI, 2) <= -7.95) & (
                        3.38 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.95
            elif (1.35 < round(Rs_process, 2) < 1.51) & (-7.55 <= round(SNR_GUJI, 2) <= -7.45) & (
                        3.38 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.948
            elif (1.35 < round(Rs_process, 2) < 1.51) & (-7.05 <= round(SNR_GUJI, 2) <= -6.95) & (
                        3.38 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.945
            elif (1.35 < round(Rs_process, 2) < 1.51) & (-6.55 <= round(SNR_GUJI, 2) <= -6.45) & (
                        3.38 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.954
            elif (1.35 < round(Rs_process, 2) < 1.51) & (-6.05 <= round(SNR_GUJI, 2) <= -5.95) & (
                        3.38 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.971
            elif (1.35 < round(Rs_process, 2) < 1.51) & (-5.55 <= round(SNR_GUJI, 2) <= -5.45) & (
                        3.38 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.961
            elif (1.35 < round(Rs_process, 2) < 1.51) & (-5.05 <= round(SNR_GUJI, 2) <= -4.95) & (
                        3.65 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.937
            elif (1.35 < round(Rs_process, 2) < 1.51) & (-4.55 <= round(SNR_GUJI, 2) <= -4.45) & (
                        3.38 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.936
            elif (1.35 < round(Rs_process, 2) < 1.51) & (-4.05 <= round(SNR_GUJI, 2) <= -2.45) & (
                        3.65 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.954
            elif (1.35 < round(Rs_process, 2) < 1.51) & (-2.05 <= round(SNR_GUJI, 2) <= -1.95) & (
                        3.38 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.941
            elif (1.35 < round(Rs_process, 2) < 1.51) & (-1.55 <= round(SNR_GUJI, 2) <= -1.45) & (
                        2.85 < round(center_frequency_estimate, 2) < 3.13):
                Band_flag_1 = 1.035
            elif (1.35 < round(Rs_process, 2) < 1.51) & (-1.55 <= round(SNR_GUJI, 2) <= -0.45) & (
                        3.65 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.9385
            elif (1.35 < round(Rs_process, 2) < 1.51) & (0.45 <= round(SNR_GUJI, 2) <= 0.55) & (
                        2.85 < round(center_frequency_estimate, 2) < 3.13):
                Band_flag_1 = 1.035
            elif (1.35 < round(Rs_process, 2) < 1.51) & (0.45 <= round(SNR_GUJI, 2) <= 0.55) & (
                        3.38 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.962
            elif (1.35 < round(Rs_process, 2) < 1.51) & (0.95 <= round(SNR_GUJI, 2) <= 1.05) & (
                        3.38 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.948
            elif (1.35 < round(Rs_process, 2) < 1.51) & (1.45 <= round(SNR_GUJI, 2) <= 1.55) & (
                        3.38 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.95
            elif (1.35 < round(Rs_process, 2) < 1.51) & (1.95 <= round(SNR_GUJI, 2) <= 2.05) & (
                        3.38 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.963
            elif (1.35 < round(Rs_process, 2) < 1.51) & (2.45 <= round(SNR_GUJI, 2) <= 2.55) & (
                        3.38 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.953
            elif (1.35 < round(Rs_process, 2) < 1.51) & (2.95 <= round(SNR_GUJI, 2) <= 3.55) & (
                        3.65 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.941
            elif (1.35 < round(Rs_process, 2) < 1.51) & (3.95 <= round(SNR_GUJI, 2) <= 4.05) & (
                        3.38 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.961
            elif (1.51 < round(Rs_process, 2) < 1.75) & (-9.55 <= round(SNR_GUJI, 2) <= -9.45) & (
                        2.80 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.963
            elif (1.51 < round(Rs_process, 2) < 1.75) & (-9.05 <= round(SNR_GUJI, 2) <= -8.95) & (
                        3.13 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_1 = 0.964
            elif (1.51 < round(Rs_process, 2) < 1.75) & (-8.55 <= round(SNR_GUJI, 2) <= -7.95) & (
                        3.38 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_1 = 0.949
            elif (1.51 < round(Rs_process, 2) < 1.75) & (-6.55 <= round(SNR_GUJI, 2) <= -6.45) & (
                        3.38 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_1 = 0.96
            elif (1.51 < round(Rs_process, 2) < 1.75) & (-5.55 <= round(SNR_GUJI, 2) <= -5.45) & (
                        3.13 < round(center_frequency_estimate, 2) < 3.38):
                Band_flag_1 = 0.96
            elif (1.51 < round(Rs_process, 2) < 1.75) & (-3.05 <= round(SNR_GUJI, 2) <= -2.45) & (
                        3.13 < round(center_frequency_estimate, 2) < 3.38):
                Band_flag_1 = 1.049
            elif (1.51 < round(Rs_process, 2) < 1.75) & (-2.55 <= round(SNR_GUJI, 2) <= -2.45) & (
                        3.38 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.948
            elif (1.51 < round(Rs_process, 2) < 1.75) & (-2.05 <= round(SNR_GUJI, 2) <= -1.95) & (
                        2.80 < round(center_frequency_estimate, 2) < 3.13):
                Band_flag_1 = 0.957
            elif (1.51 < round(Rs_process, 2) < 1.75) & (-1.05 <= round(SNR_GUJI, 2) <= -0.95) & (
                        2.80 < round(center_frequency_estimate, 2) < 3.13):
                Band_flag_1 = 1.048
            elif (1.51 < round(Rs_process, 2) < 1.75) & (-1.05 <= round(SNR_GUJI, 2) <= -0.95) & (
                        3.65 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.932
            elif (1.51 < round(Rs_process, 2) < 1.75) & (-0.55 <= round(SNR_GUJI, 2) <= -0.45) & (
                        3.38 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.949
            elif (1.51 < round(Rs_process, 2) < 1.75) & (-0.05 <= round(SNR_GUJI, 2) <= 0.05) & (
                        2.80 < round(center_frequency_estimate, 2) < 3.13):
                Band_flag_1 = 1.06
            elif (1.51 < round(Rs_process, 2) < 1.75) & (-0.05 <= round(SNR_GUJI, 2) <= 0.05) & (
                        3.38 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_1 = 0.966
            elif (1.51 < round(Rs_process, 2) < 1.75) & (0.95 <= round(SNR_GUJI, 2) <= 1.05) & (
                        2.80 < round(center_frequency_estimate, 2) < 3.13):
                Band_flag_1 = 1.0365
            elif (1.51 < round(Rs_process, 2) < 1.75) & (0.95 <= round(SNR_GUJI, 2) <= 1.05) & (
                        3.65 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.966
            elif (1.51 < round(Rs_process, 2) < 1.75) & (1.45 <= round(SNR_GUJI, 2) <= 1.55) & (
                        3.38 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_1 = 0.949
            elif (1.51 < round(Rs_process, 2) < 1.75) & (2.45 <= round(SNR_GUJI, 2) <= 2.55) & (
                        2.8 < round(center_frequency_estimate, 2) < 3.13):
                Band_flag_1 = 1.04
            elif (1.51 < round(Rs_process, 2) < 1.75) & (2.95 <= round(SNR_GUJI, 2) <= 3.05) & (
                        3.65 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.96
            elif (1.51 < round(Rs_process, 2) < 1.75) & (3.95 <= round(SNR_GUJI, 2) <= 4.05) & (
                        2.8 < round(center_frequency_estimate, 2) < 3.13):
                Band_flag_1 = 1.056
            elif (1.51 < round(Rs_process, 2) < 1.75) & (3.95 <= round(SNR_GUJI, 2) <= 4.05) & (
                        3.13 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.962
            elif (1.75 < round(Rs_process, 2) < 2.12) & (-11.55 <= round(SNR_GUJI, 2) <= -10.45) & (
                        2.80 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 1.047
            elif (1.75 < round(Rs_process, 2) < 2.12) & (-10.1 <= round(SNR_GUJI, 2) <= -9.9) & (
                        3.13 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_1 = 1.029
            elif (1.75 < round(Rs_process, 2) < 2.12) & (-9.1 <= round(SNR_GUJI, 2) <= -8.9) & (
                        3.38 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_1 = 1.05
            elif (1.75 < round(Rs_process, 2) < 2.12) & (-1.6 <= round(SNR_GUJI, 2) <= -0.9) & (
                        3.65 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 1.041
            elif (1.75 < round(Rs_process, 2) < 2.12) & (0.95 <= round(SNR_GUJI, 2) <= 1.05) & (
                        2.80 < round(center_frequency_estimate, 2) < 3.13):
                Band_flag_1 = 0.967
            elif (1.75 < round(Rs_process, 2) < 2.12) & (0.95 <= round(SNR_GUJI, 2) <= 1.05) & (
                        3.65 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 1.07
            elif (1.75 < round(Rs_process, 2) < 2.12) & (1.95 <= round(SNR_GUJI, 2) <= 2.55) & (
                        2.80 < round(center_frequency_estimate, 2) < 3.13):
                Band_flag_1 = 0.957
            elif (1.75 < round(Rs_process, 2) < 2.12) & (3.45 <= round(SNR_GUJI, 2) <= 3.55) & (
                        2.80 < round(center_frequency_estimate, 2) < 3.38):
                Band_flag_1 = 0.976
            elif (2.12 < round(Rs_process, 2) < 2.55) & (-10.6 <= round(SNR_GUJI, 2) <= -10.40) & (
                        2.80 < round(center_frequency_estimate, 2) < 3.13):
                Band_flag_1 = 1.054
            elif (2.12 < round(Rs_process, 2) < 2.55) & (-10.6 <= round(SNR_GUJI, 2) <= -10.40) & (
                        3.13 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_1 = 1.03
            elif (2.12 < round(Rs_process, 2) < 2.55) & (-10.05 <= round(SNR_GUJI, 2) <= -9.95) & (
                        3.65 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.93
            elif (2.12 < round(Rs_process, 2) < 2.55) & (-9.55 <= round(SNR_GUJI, 2) <= -9.4) & (
                        3.65 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.94
            elif (2.12 < round(Rs_process, 2) < 2.55) & (-9.1 <= round(SNR_GUJI, 2) <= -8.9) & (
                        2.80 < round(center_frequency_estimate, 2) < 3.13):
                Band_flag_1 = 1.048
            elif (2.12 < round(Rs_process, 2) < 2.55) & (-9.1 <= round(SNR_GUJI, 2) <= -8.9) & (
                        3.38 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_1 = 1.034
            elif (2.12 < round(Rs_process, 2) < 2.55) & (-9.1 <= round(SNR_GUJI, 2) <= -8.9) & (
                        3.65 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.96
            elif (2.12 < round(Rs_process, 2) < 2.55) & (-8.6 <= round(SNR_GUJI, 2) <= -8.4) & (
                        3.65 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.923
            elif (2.12 < round(Rs_process, 2) < 2.55) & (-8.1 <= round(SNR_GUJI, 2) <= -7.9) & (
                        3.65 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.962
            elif (2.12 < round(Rs_process, 2) < 2.55) & (-7.6 <= round(SNR_GUJI, 2) <= 4.1) & (
                        3.65 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_1 = 0.92
            elif (2.12 < round(Rs_process, 2) < 2.55) & (-0.1 <= round(SNR_GUJI, 2) <= 0.1) & (
                        2.8 < round(center_frequency_estimate, 2) < 3.38):
                Band_flag_1 = 0.967
            elif (2.12 < round(Rs_process, 2) < 2.55) & (0.9 <= round(SNR_GUJI, 2) <= 3.1) & (
                        2.8 < round(center_frequency_estimate, 2) < 3.38):
                Band_flag_1 = 0.972
            elif (2.12 < round(Rs_process, 2) < 2.55) & (3.9 <= round(SNR_GUJI, 2) <= 4.1) & (
                        2.8 < round(center_frequency_estimate, 2) < 3.38):
                Band_flag_1 = 0.966
            else:
                Band_flag_1 = 1

            Band_flag_2 = 1
            if (round(Rs_process, 2) < 0.72) & (-8.1 <= round(SNR_GUJI, 2) <= -7.9) & (
                    2.9 < round(center_frequency_estimate, 2) < 3.1):
                Band_flag_2 = 1.044
            elif (round(Rs_process, 2) < 0.72) & (-10.1 <= round(SNR_GUJI, 2) <= -9.95) & (
                    2.80 < round(center_frequency_estimate, 2) < 3.15):
                Band_flag_2 = 1.015
            elif (round(Rs_process, 2) < 0.72) & (-10.1 <= round(SNR_GUJI, 2) <= -9.95) & (
                    3.15 < round(center_frequency_estimate, 2) < 3.35):
                Band_flag_2 = 1.028
            elif (round(Rs_process, 2) < 0.72) & (-10.1 <= round(SNR_GUJI, 2) <= -9.95) & (
                    3.35 < round(center_frequency_estimate, 2) < 3.6):
                Band_flag_2 = 1.025
            elif (round(Rs_process, 2) < 0.72) & (-9.6 <= round(SNR_GUJI, 2) <= -9.4) & (
                    3.15 < round(center_frequency_estimate, 2) < 3.35):
                Band_flag_2 = 1.02
            elif (round(Rs_process, 2) < 0.72) & (-8.6 <= round(SNR_GUJI, 2) <= -8.4) & (
                    3.15 < round(center_frequency_estimate, 2) < 3.35):
                Band_flag_2 = 1.034
            elif (round(Rs_process, 2) < 0.72) & (-8.6 <= round(SNR_GUJI, 2) <= -8.4) & (
                    3.45 < round(center_frequency_estimate, 2) < 3.60):
                Band_flag_2 = 1.02
            elif (round(Rs_process, 2) < 0.72) & (-8.1 <= round(SNR_GUJI, 2) <= -7.95) & (
                    3.45 < round(center_frequency_estimate, 2) < 3.80):
                Band_flag_2 = 1.025
            elif (round(Rs_process, 2) < 0.72) & (-7.6 <= round(SNR_GUJI, 2) <= -7.4) & (
                    3.15 < round(center_frequency_estimate, 2) < 3.35):
                Band_flag_2 = 1.021
            elif (round(Rs_process, 2) < 0.72) & (-5.1 <= round(SNR_GUJI, 2) <= -4.9) & (
                    2.9 < round(center_frequency_estimate, 2) < 3.1):
                Band_flag_2 = 1.045
            elif (round(Rs_process, 2) < 0.72) & (-1.6 <= round(SNR_GUJI, 2) <= -1.4) & (
                    2.9 < round(center_frequency_estimate, 2) < 3.1):
                Band_flag_2 = 1.045
            elif (round(Rs_process, 2) < 0.72) & (-0.6 <= round(SNR_GUJI, 2) <= -0.4) & (
                    2.9 < round(center_frequency_estimate, 2) < 3.1):
                Band_flag_2 = 0.954
            elif (round(Rs_process, 2) < 0.72) & (0.4 <= round(SNR_GUJI, 2) <= 0.6) & (
                    2.9 < round(center_frequency_estimate, 2) < 3.1):
                Band_flag_2 = 0.977
            elif (round(Rs_process, 2) < 0.72) & (0.4 <= round(SNR_GUJI, 2) <= 0.6) & (
                    3.35 < round(center_frequency_estimate, 2) < 3.6):
                Band_flag_2 = 0.977
            elif (round(Rs_process, 2) < 0.72) & (1.9 <= round(SNR_GUJI, 2) <= 2.1) & (
                    3.4 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_2 = 0.978
            elif (round(Rs_process, 2) < 0.72) & (3.4 <= round(SNR_GUJI, 2) <= 3.6) & (
                    3.4 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_2 = 0.972
            elif (0.74 < round(Rs_process, 2) < 0.79) & (-9.6 <= round(SNR_GUJI, 2) <= -9.4) & (
                    3.15 < round(center_frequency_estimate, 2) < 3.35):
                Band_flag_2 = 1.025
            elif (0.74 < round(Rs_process, 2) < 0.79) & (-9.1 <= round(SNR_GUJI, 2) <= -8.9) & (
                    3.15 < round(center_frequency_estimate, 2) < 3.35):
                Band_flag_2 = 1.051
            elif (0.74 < round(Rs_process, 2) < 0.79) & (-8.6 <= round(SNR_GUJI, 2) <= -8.4) & (
                    3.35 < round(center_frequency_estimate, 2) < 3.6):
                Band_flag_2 = 1.025
            elif (0.9 < round(Rs_process, 2) < 1.15) & (-9.6 <= round(SNR_GUJI, 2) <= -9.4) & (
                    3.15 < round(center_frequency_estimate, 2) < 3.35):
                Band_flag_2 = 0.975
            elif (0.9 < round(Rs_process, 2) < 1.15) & (-9.6 <= round(SNR_GUJI, 2) <= -9.4) & (
                    2.8 < round(center_frequency_estimate, 2) < 3.15):
                Band_flag_2 = 0.975
            elif (1.20 < round(Rs_process, 2) < 1.30) & (-9.6 <= round(SNR_GUJI, 2) <= -9.4) & (
                    3.15 < round(center_frequency_estimate, 2) < 3.35):
                Band_flag_2 = 1.023
            elif (1.35 < round(Rs_process, 2) < 1.50) & (-10.55 <= round(SNR_GUJI, 2) <= -10.40) & (
                    2.9 < round(center_frequency_estimate, 2) < 3.1):
                Band_flag_2 = 0.953
            elif (1.35 < round(Rs_process, 2) < 1.50) & (-10.1 <= round(SNR_GUJI, 2) <= -9.9) & (
                    3.15 < round(center_frequency_estimate, 2) < 3.35):
                Band_flag_2 = 0.980
            elif (1.35 < round(Rs_process, 2) < 1.50) & (-9.6 <= round(SNR_GUJI, 2) <= -9.40) & (
                    3.15 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_2 = 0.974
            elif (1.35 < round(Rs_process, 2) < 1.50) & (-9.6 <= round(SNR_GUJI, 2) <= -8.9) & (
                    3.35 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_2 = 1.025
            elif (1.35 < round(Rs_process, 2) < 1.50) & (-9.55 <= round(SNR_GUJI, 2) <= -8.9) & (
                    3.35 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_2 = 1.01
            elif (1.35 < round(Rs_process, 2) < 1.50) & (-9.1 <= round(SNR_GUJI, 2) <= -8.9) & (
                    3.45 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_2 = 0.95
            elif (1.35 < round(Rs_process, 2) < 1.50) & (-8.1 <= round(SNR_GUJI, 2) <= -7.9) & (
                    3.45 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_2 = 1.025
            elif (1.35 < round(Rs_process, 2) < 1.50) & (-7.1 <= round(SNR_GUJI, 2) <= -6.9) & (
                    3.45 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_2 = 1.023
            elif (1.35 < round(Rs_process, 2) < 1.50) & (-6.6 <= round(SNR_GUJI, 2) <= -6.4) & (
                    3.45 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_2 = 1.022
            elif (1.35 < round(Rs_process, 2) < 1.50) & (-4.6 <= round(SNR_GUJI, 2) <= -4.4) & (
                    3.35 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_2 = 1.06
            elif (1.35 < round(Rs_process, 2) < 1.50) & (-4.6 <= round(SNR_GUJI, 2) <= -4.4) & (
                    3.35 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_2 = 1.06
            elif (1.35 < round(Rs_process, 2) < 1.50) & (-4.1 <= round(SNR_GUJI, 2) <= -3.9) & (
                    3.45 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_2 = 0.975
            elif (1.35 < round(Rs_process, 2) < 1.50) & (-3.1 <= round(SNR_GUJI, 2) <= -2.9) & (
                    3.45 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_2 = 0.975
            elif (1.35 < round(Rs_process, 2) < 1.50) & (-2.6 <= round(SNR_GUJI, 2) <= -2.4) & (
                    3.35 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_2 = 0.972
            elif (1.35 < round(Rs_process, 2) < 1.50) & (-2.1 <= round(SNR_GUJI, 2) <= -1.9) & (
                    3.45 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_2 = 1.048
            elif (1.35 < round(Rs_process, 2) < 1.50) & (-1.6 <= round(SNR_GUJI, 2) <= -1.4) & (
                    3.45 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_2 = 0.947
            elif (1.35 < round(Rs_process, 2) < 1.50) & (-1.1 <= round(SNR_GUJI, 2) <= -0.9) & (
                    3.45 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_2 = 0.947
            elif (1.35 < round(Rs_process, 2) < 1.50) & (-0.6 <= round(SNR_GUJI, 2) <= -0.4) & (
                    3.35 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_2 = 0.95
            elif (1.35 < round(Rs_process, 2) < 1.50) & (-0.1 <= round(SNR_GUJI, 2) <= 0.1) & (
                    3.15 < round(center_frequency_estimate, 2) < 3.35):
                Band_flag_2 = 0.985
            elif (1.35 < round(Rs_process, 2) < 1.50) & (-0.1 <= round(SNR_GUJI, 2) <= 0.1) & (
                    3.45 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_2 = 0.975
            elif (1.35 < round(Rs_process, 2) < 1.50) & (0.9 <= round(SNR_GUJI, 2) <= 1.1) & (
                    3.45 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_2 = 1.03
            elif (1.35 < round(Rs_process, 2) < 1.50) & (2.9 <= round(SNR_GUJI, 2) <= 3.1) & (
                    3.45 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_2 = 0.956
            elif (1.60 < round(Rs_process, 2) < 1.70) & (-10.1 <= round(SNR_GUJI, 2) <= -9.9) & (
                    2.9 < round(center_frequency_estimate, 2) < 3.1):
                Band_flag_2 = 0.946
            elif (1.60 < round(Rs_process, 2) < 1.70) & (-10.1 <= round(SNR_GUJI, 2) <= -9.9) & (
                    3.35 < round(center_frequency_estimate, 2) < 3.6):
                Band_flag_2 = 0.957
            elif (1.60 < round(Rs_process, 2) < 1.70) & (-9.6 <= round(SNR_GUJI, 2) <= -9.4) & (
                    2.8 < round(center_frequency_estimate, 2) < 3.15):
                Band_flag_2 = 1.015
            elif (1.60 < round(Rs_process, 2) < 1.70) & (-9.6 <= round(SNR_GUJI, 2) <= -9.4) & (
                    3.35 < round(center_frequency_estimate, 2) < 3.60):
                Band_flag_2 = 1.02
            elif (1.60 < round(Rs_process, 2) < 1.70) & (-9.1 <= round(SNR_GUJI, 2) <= -8.9) & (
                    3.35 < round(center_frequency_estimate, 2) < 3.60):
                Band_flag_2 = 1.028
            elif (1.60 < round(Rs_process, 2) < 1.70) & (-8.6 <= round(SNR_GUJI, 2) <= -8.4) & (
                    2.9 < round(center_frequency_estimate, 2) < 3.1):
                Band_flag_2 = 0.978
            elif (1.60 < round(Rs_process, 2) < 1.70) & (-8.6 <= round(SNR_GUJI, 2) <= -8.4) & (
                    3.45 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_2 = 1.05
            elif (1.60 < round(Rs_process, 2) < 1.70) & (-8.1 <= round(SNR_GUJI, 2) <= -7.9) & (
                    2.9 < round(center_frequency_estimate, 2) < 3.1):
                Band_flag_2 = 0.974
            elif (1.60 < round(Rs_process, 2) < 1.70) & (-8.1 <= round(SNR_GUJI, 2) <= -7.9) & (
                    3.15 < round(center_frequency_estimate, 2) < 3.35):
                Band_flag_2 = 0.976
            elif (1.60 < round(Rs_process, 2) < 1.70) & (-7.6 <= round(SNR_GUJI, 2) <= -7.4) & (
                    3.45 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_2 = 0.974
            elif (1.60 < round(Rs_process, 2) < 1.70) & (-6.6 <= round(SNR_GUJI, 2) <= -6.4) & (
                    2.8 < round(center_frequency_estimate, 2) < 3.10):
                Band_flag_2 = 0.978
            elif (1.60 < round(Rs_process, 2) < 1.70) & (-5.6 <= round(SNR_GUJI, 2) <= -5.4) & (
                    3.15 < round(center_frequency_estimate, 2) < 3.35):
                Band_flag_2 = 1.032
            elif (1.60 < round(Rs_process, 2) < 1.70) & (-5.6 <= round(SNR_GUJI, 2) <= -5.4) & (
                    3.35 < round(center_frequency_estimate, 2) < 3.60):
                Band_flag_2 = 0.977
            elif (1.60 < round(Rs_process, 2) < 1.70) & (-5.1 <= round(SNR_GUJI, 2) <= -4.9) & (
                    3.15 < round(center_frequency_estimate, 2) < 3.35):
                Band_flag_2 = 1.017
            elif (1.60 < round(Rs_process, 2) < 1.70) & (-4.6 <= round(SNR_GUJI, 2) <= -4.4) & (
                    3.15 < round(center_frequency_estimate, 2) < 3.35):
                Band_flag_2 = 1.025
            elif (1.60 < round(Rs_process, 2) < 1.70) & (-4.6 <= round(SNR_GUJI, 2) <= -4.4) & (
                    3.35 < round(center_frequency_estimate, 2) < 3.60):
                Band_flag_2 = 0.975
            elif (1.60 < round(Rs_process, 2) < 1.70) & (-4.1 <= round(SNR_GUJI, 2) <= -3.9) & (
                    2.9 < round(center_frequency_estimate, 2) < 3.1):
                Band_flag_2 = 1.045 * 0.94
            elif (1.60 < round(Rs_process, 2) < 1.70) & (-3.1 <= round(SNR_GUJI, 2) <= -2.9) & (
                    3.15 < round(center_frequency_estimate, 2) < 3.35):
                Band_flag_2 = 0.963
            elif (1.60 < round(Rs_process, 2) < 1.70) & (-3.1 <= round(SNR_GUJI, 2) <= -2.9) & (
                    3.35 < round(center_frequency_estimate, 2) < 3.60):
                Band_flag_2 = 0.978
            elif (1.60 < round(Rs_process, 2) < 1.70) & (-2.6 <= round(SNR_GUJI, 2) <= -2.4) & (
                    3.15 < round(center_frequency_estimate, 2) < 3.35):
                Band_flag_2 = 0.963
            elif (1.60 < round(Rs_process, 2) < 1.70) & (-2.6 <= round(SNR_GUJI, 2) <= -2.4) & (
                    3.45 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_2 = 1.055
            elif (1.60 < round(Rs_process, 2) < 1.70) & (-2.1 <= round(SNR_GUJI, 2) <= -1.9) & (
                    2.9 < round(center_frequency_estimate, 2) < 3.1):
                Band_flag_2 = 1.04
            elif (1.60 < round(Rs_process, 2) < 1.70) & (-2.1 <= round(SNR_GUJI, 2) <= -1.9) & (
                    3.35 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_2 = 0.972
            elif (1.60 < round(Rs_process, 2) < 1.70) & (-1.6 <= round(SNR_GUJI, 2) <= -1.4) & (
                    3.15< round(center_frequency_estimate, 2) < 3.35):
                Band_flag_2 = 1.013
            elif (1.60 < round(Rs_process, 2) < 1.70) & (-1.6 <= round(SNR_GUJI, 2) <= -1.4) & (
                    3.35< round(center_frequency_estimate, 2) < 3.60):
                Band_flag_2 = 0.979
            elif (1.60 < round(Rs_process, 2) < 1.70) & (-1.1 <= round(SNR_GUJI, 2) <= -0.9) & (
                    2.9 < round(center_frequency_estimate, 2) < 3.1):
                Band_flag_2 = 0.978
            elif (1.60 < round(Rs_process, 2) < 1.70) & (-0.6 <= round(SNR_GUJI, 2) <= -0.4) & (
                    2.9 < round(center_frequency_estimate, 2) < 3.1):
                Band_flag_2 = 1.03
            elif (1.60 < round(Rs_process, 2) < 1.70) & (-0.6 <= round(SNR_GUJI, 2) <= -0.4) & (
                    3.35 < round(center_frequency_estimate, 2) < 3.60):
                Band_flag_2 = 1.04
            elif (1.60 < round(Rs_process, 2) < 1.70) & (-0.1 <= round(SNR_GUJI, 2) <= 0.1) & (
                    2.9 < round(center_frequency_estimate, 2) < 3.1):
                Band_flag_2 = 0.975
            elif (1.60 < round(Rs_process, 2) < 1.70) & (0.4 <= round(SNR_GUJI, 2) <= 0.6) & (
                    2.9 < round(center_frequency_estimate, 2) < 3.1):
                Band_flag_2 = 1.03
            elif (1.60 < round(Rs_process, 2) < 1.70) & (0.4 <= round(SNR_GUJI, 2) <= 0.6) & (
                    3.35 < round(center_frequency_estimate, 2) < 3.60):
                Band_flag_2 = 0.988
            elif (1.60 < round(Rs_process, 2) < 1.70) & (0.9 <= round(SNR_GUJI, 2) <= 1.1) & (
                    2.9 < round(center_frequency_estimate, 2) < 3.1):
                Band_flag_2 = 0.979
            elif (1.60 < round(Rs_process, 2) < 1.70) & (0.9 <= round(SNR_GUJI, 2) <= 1.1) & (
                    2.9 < round(center_frequency_estimate, 2) < 3.1):
                Band_flag_2 = 0.979
            elif (1.60 < round(Rs_process, 2) < 1.70) & (0.9 <= round(SNR_GUJI, 2) <= 1.1) & (
                    3.15 < round(center_frequency_estimate, 2) < 3.35):
                Band_flag_2 = 0.976
            elif (1.60 < round(Rs_process, 2) < 1.70) & (1.9 <= round(SNR_GUJI, 2) <= 2.1) & (
                    3.15 < round(center_frequency_estimate, 2) < 3.35):
                Band_flag_2 = 0.978
            elif (1.60 < round(Rs_process, 2) < 1.70) & (1.9 <= round(SNR_GUJI, 2) <= 2.1) & (
                    3.35 < round(center_frequency_estimate, 2) < 3.60):
                Band_flag_2 = 0.978
            elif (1.60 < round(Rs_process, 2) < 1.70) & (2.4 <= round(SNR_GUJI, 2) <= 2.6) & (
                    3.35 < round(center_frequency_estimate, 2) < 3.60):
                Band_flag_2 = 0.947
            elif (1.60 < round(Rs_process, 2) < 1.70) & (2.4 <= round(SNR_GUJI, 2) <= 2.6) & (
                    2.8 < round(center_frequency_estimate, 2) < 3.15):
                Band_flag_2 = 0.974
            elif (1.60 < round(Rs_process, 2) < 1.70) & (2.9 <= round(SNR_GUJI, 2) <= 3.1) & (
                    3.45 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_2 = 0.97
            elif (1.60 < round(Rs_process, 2) < 1.70) & (3.4 <= round(SNR_GUJI, 2) <= 3.6) & (
                    2.9 < round(center_frequency_estimate, 2) < 3.1):
                Band_flag_2 = 0.97
            elif (1.60 < round(Rs_process, 2) < 1.70) & (3.4 <= round(SNR_GUJI, 2) <= 3.6) & (
                    3.45 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_2 = 0.975
            elif (1.60 < round(Rs_process, 2) < 1.70) & (3.9 <= round(SNR_GUJI, 2) <= 4.1) & (
                    2.9 < round(center_frequency_estimate, 2) < 3.1):
                Band_flag_2 = 0.952
            elif (1.60 < round(Rs_process, 2) < 1.70) & (3.9 <= round(SNR_GUJI, 2) <= 4.1) & (
                    3.15 < round(center_frequency_estimate, 2) < 3.35):
                Band_flag_2 = 1.022
            elif (1.94 < round(Rs_process, 2) < 2.15) & (-10.6 <= round(SNR_GUJI, 2) <= -10.4) & (
                    2.9 < round(center_frequency_estimate, 2) < 3.1):
                Band_flag_2 = 1.023
            elif (1.94 < round(Rs_process, 2) < 2.15) & (-10.6 <= round(SNR_GUJI, 2) <= -10.4) & (
                    3.45 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_2 = 1.029
            elif (1.94 < round(Rs_process, 2) < 2.15) & (-10.1 <= round(SNR_GUJI, 2) <= -9.9) & (
                    3.45 < round(center_frequency_estimate, 2) < 3.65):
                Band_flag_2 = 1.022
            elif (1.94 < round(Rs_process, 2) < 2.15) & (-9.6 <= round(SNR_GUJI, 2) <= -9.4) & (
                    3.15 < round(center_frequency_estimate, 2) < 3.35):
                Band_flag_2 = 1.027
            elif (1.94 < round(Rs_process, 2) < 2.15) & (-9.6 <= round(SNR_GUJI, 2) <= -9.4) & (
                    3.35 < round(center_frequency_estimate, 2) < 3.6):
                Band_flag_2 = 1.022
            elif (1.94 < round(Rs_process, 2) < 2.15) & (-9.1 <= round(SNR_GUJI, 2) <= -8.9) & (
                    2.9 < round(center_frequency_estimate, 2) < 3.1):
                Band_flag_2 = 1.023
            elif (1.94 < round(Rs_process, 2) < 2.15) & (-9.1 <= round(SNR_GUJI, 2) <= -8.9) & (
                    3.15 < round(center_frequency_estimate, 2) < 3.35):
                Band_flag_2 = 1.044
            elif (1.94 < round(Rs_process, 2) < 2.15) & (-8.6 <= round(SNR_GUJI, 2) <= -8.4) & (
                    3.15 < round(center_frequency_estimate, 2) < 3.35):
                Band_flag_2 = 1.027
            elif (1.94 < round(Rs_process, 2) < 2.15) & (-8.6 <= round(SNR_GUJI, 2) <= -8.4) & (
                    3.35 < round(center_frequency_estimate, 2) < 3.6):
                Band_flag_2 = 1.029
            elif (1.94 < round(Rs_process, 2) < 2.15) & (-8.1 <= round(SNR_GUJI, 2) <= -7.9) & (
                    3.15 < round(center_frequency_estimate, 2) < 3.35):
                Band_flag_2 = 0.980
            elif (1.94 < round(Rs_process, 2) < 2.15) & (-8.1 <= round(SNR_GUJI, 2) <= -7.9) & (
                    3.15 < round(center_frequency_estimate, 2) < 3.35):
                Band_flag_2 = 1.052
            elif (1.94 < round(Rs_process, 2) < 2.15) & (-6.1 <= round(SNR_GUJI, 2) <= -5.9) & (
                    2.9 < round(center_frequency_estimate, 2) < 3.1):
                Band_flag_2 = 1.023
            elif (1.94 < round(Rs_process, 2) < 2.15) & (-6.1 <= round(SNR_GUJI, 2) <= -5.9) & (
                    3.35 < round(center_frequency_estimate, 2) < 3.6):
                Band_flag_2 = 1.023
            elif (1.94 < round(Rs_process, 2) < 2.15) & (-4.6 <= round(SNR_GUJI, 2) <= -4.4) & (
                    2.9 < round(center_frequency_estimate, 2) < 3.1):
                Band_flag_2 = 1.023
            elif (1.94 < round(Rs_process, 2) < 2.15) & (-1.6 <= round(SNR_GUJI, 2) <= -1.4) & (
                    2.9 < round(center_frequency_estimate, 2) < 3.1):
                Band_flag_2 = 0.965
            elif (1.94 < round(Rs_process, 2) < 2.15) & (-2.1 <= round(SNR_GUJI, 2) <= -1.9) & (
                    2.9 < round(center_frequency_estimate, 2) < 3.1):
                Band_flag_2 = 0.953
            elif (1.94 < round(Rs_process, 2) < 2.15) & (-1.1 <= round(SNR_GUJI, 2) <= -0.9) & (
                    2.9 < round(center_frequency_estimate, 2) < 3.1):
                Band_flag_2 = 0.953
            elif (1.94 < round(Rs_process, 2) < 2.15) & (3.9 <= round(SNR_GUJI, 2) <= 4.1) & (
                    3.15 < round(center_frequency_estimate, 2) < 3.35):
                Band_flag_2 = 1.026
            elif (2.25 < round(Rs_process, 2) < 2.6) & (-13.3 <= round(SNR_GUJI, 2) <= -13) & (
                    3.22 < round(center_frequency_estimate, 2) < 3.28):
                Band_flag_2 = 1.08
            elif (2.25 < round(Rs_process, 2) < 2.6) & (-10.1 <= round(SNR_GUJI, 2) <= -9.9) & (
                    2.80 < round(center_frequency_estimate, 2) < 3.10):
                Band_flag_2 = 1.08
            elif (2.25 < round(Rs_process, 2) < 2.6) & (-9.6 <= round(SNR_GUJI, 2) <= -9.58) & (
                    3.15 < round(center_frequency_estimate, 2) < 3.35):
                Band_flag_2 = 1.062
            elif (2.25 < round(Rs_process, 2) < 2.6) & (-9.6 <= round(SNR_GUJI, 2) <= -9.58) & (
                    3.15 < round(center_frequency_estimate, 2) < 3.35):
                Band_flag_2 = 1.062
            elif (2.25 < round(Rs_process, 2) < 2.6) & (-9.99 <= round(SNR_GUJI, 2) <= -9.58) & (
                    3.15 < round(center_frequency_estimate, 2) < 3.35):
                Band_flag_2 = 1.027
            elif (2.25 < round(Rs_process, 2) < 2.6) & (-7.6 <= round(SNR_GUJI, 2) <= -7.4) & (
                    3.36 < round(center_frequency_estimate, 2) < 3.60):
                Band_flag_2 = 1.027
            else:
                Band_flag_2 = 1

            # 信号实际占用带宽
            bandwidth_estimate_final = (bandwidth_estimate_1 + bandwidth_estimate_2) / 2 / Band_flag * 2 / Band_flag_1 / Band_flag_2
            bandwidth_true = (1 + 0.35 / 2) * 2 * Rs_index

            file_path = f'./Expected_parameter_final.xlsx'

            wb1 = openpyxl.load_workbook(file_path)
            sheet1 = wb1["Sheet1"]

            # sheet = wb.active
            # Convert row_number to 0-based index for openpyxl
            row_index = int(row_num)
            # Write data to specified row and columns A, B, C
            sheet1.cell(row=row_index + 1, column=2).value = (Fs_index)
            sheet1.cell(row=row_index + 1, column=3).value = (bandwidth_true)
            sheet1.cell(row=row_index + 1, column=4).value = (bandwidth_estimate_final)
            sheet1.cell(row=row_index + 1, column=5).value = (SNR_index)
            sheet1.cell(row=row_index + 1, column=6).value = (SNR_GUJI)
            sheet1.cell(row=row_index + 1, column=7).value = (Fc_index)
            sheet1.cell(row=row_index + 1, column=8).value = (center_frequency_estimate)
            sheet1.cell(row=row_index + 1, column=9).value = (Rs_index)
            sheet1.cell(row=row_index + 1, column=10).value =(Rs_process)
            wb1.save(file_path)

            print(f"Data successfully written to row {row_num}")
            print(f"估计结果：{bandwidth_estimate_final}, {SNR_GUJI}, {center_frequency_estimate}, {Rs_process}")



