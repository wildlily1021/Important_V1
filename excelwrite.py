import openpyxl
import psutil
import os
from PyQt5.QtWidgets import QMessageBox


# def close_excel_file(file_path):
#     for proc in psutil.process_iter():
#         try:
#             for item in proc.open_files():
#                 if item.path == os.path.abspath(file_path):
#                     return 1
#                     # proc.kill()
#         except:
#             pass

def write_to_excel(row_num, data1, data2, data3):
    file_path = f'./Expected_parameter_-8.xlsx'
    # Select the active sheet (assuming there is only one sheet)
    # Load the workbook
    # close_excel_file(file_path)
    wb = openpyxl.load_workbook(file_path)
    sheet = wb["Sheet2"]
    # sheet = wb.active
    # Convert row_number to 0-based index for openpyxl
    row_index = row_num

    # Write data to specified row and columns A, B, C
    sheet.cell(row=row_index + 1, column=1).value = data1
    sheet.cell(row=row_index + 1, column=2).value = data2
    sheet.cell(row=row_index + 1, column=3).value = data3
    # Save the workbook
    wb.save(file_path)
    print(f"Data successfully written to row {row_num}")


def write_to_excel_signal(row_num, data1, data2, data3, data4, data5):#(row_num, Fs, Fc, Rs, SNR)
    file_path = f'./Expected_parameter_-8.xlsx'
    # Select the active sheet (assuming there is only one sheet)
    # Load the workbook
    # close_excel_file(file_path)
    wb = openpyxl.load_workbook(file_path)
    sheet = wb["Sheet1"]
    # sheet = wb.active
    # Convert row_number to 0-based index for openpyxl
    row_index = row_num

    # Write data to specified row and columns A, B, C
    sheet.cell(row=row_index + 1, column=1).value = row_num
    sheet.cell(row=row_index + 1, column=2).value = data1
    sheet.cell(row=row_index + 1, column=3).value = data2
    sheet.cell(row=row_index + 1, column=4).value = data3
    sheet.cell(row=row_index + 1, column=5).value = data4
    sheet.cell(row=row_index + 1, column=6).value = data5
    # Save the workbook
    wb.save(file_path)
    print(f"Data successfully written to row {row_num}")
