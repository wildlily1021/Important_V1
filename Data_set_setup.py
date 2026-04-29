import cv2
import numpy as np
import openpyxl
import tensorflow as tf
from signal_photo_save import signal_create_test
from image_processing_41 import process_image_test

for i in np.arange(1.0, 1.25, 0.25):
    for j in range(1, 11):
        for k in range(1, 101):
            Fs = 10 * 1e9
            Fs_index = Fs / 1e9
            Rs = i * 1e9
            Rs_index = Rs / 1e9
            Fc = (3.0 + j * 0.05 - 0.05) * 1e9
            Fc_index = Fc / 1e9
            SNR = -10 + (k - 1) * 0.2
            Modulation = 1
            [Fs, rec_wave, SNR_GUJI, RS_GUJI] = signal_create_test(Fs, Fc, Rs, SNR, Modulation)
            [height, center_frequency_output] = process_image_test(i, SNR, Fc_index, rec_wave, Fs, Modulation)
            center_frequency = center_frequency_output / 1.001 * 0.993
            row_num = ((i - 0.5) / 0.25) * 10 * 100 + (j - 1) * 100 + k
            # file_path = f'./Expected_parameter.xlsx'
            #
            # wb1 = openpyxl.load_workbook(file_path)
            # sheet1 = wb1["Sheet1"]
            #
            # # sheet = wb.active
            # # Convert row_number to 0-based index for openpyxl
            # row_index = int(row_num)
            # # # Write data to specified row and columns A, B, C
            # # sheet1.cell(row=row_index + 1, column=3).value = (Fc / 1e9)
            # # sheet1.cell(row=row_index + 1, column=4).value = i
            # # sheet1.cell(row=row_index + 1, column=6).value = SNR
            # # wb1.save(file_path)
            #
            # wb2 = openpyxl.load_workbook(file_path)
            # sheet2 = wb2["Sheet2"]
            # # sheet2.cell(row=row_index + 1, column=1).value = height
            # sheet2.cell(row=row_index + 1, column=2).value = (predictions_float / calculate_flag(Rs, SNR, Fs_index, Fc_index) / flag_final_1(Rs, SNR, Fs_index, Fc_index) / flag_final_2(Rs, SNR, Fs_index, Fc_index))
            # # sheet2.cell(row=row_index + 1, column=3).value = center_frequency
            # # sheet2.cell(row=row_index + 1, column=6).value = SNR_GUJI
            # # sheet2.cell(row=row_index + 1, column=7).value = (RS_GUJI / 1e9)

            # wb2.save(file_path)
            print(f"Data successfully written to row {row_num}")



