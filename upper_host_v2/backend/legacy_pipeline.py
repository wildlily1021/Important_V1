from __future__ import annotations

import base64
import contextlib
import mimetypes
import os
import re
import sys
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Any

os.environ.setdefault("MPLBACKEND", "Agg")

PROJECT_ROOT = Path(__file__).resolve().parents[2]
SIGNAL_ANA_DIR = PROJECT_ROOT / "signal_ana"

if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from stft_pipeline import (
    compute_spectrogram_matrix,
    frequency_to_pixel_y,
    load_current_stft_run,
    load_stft_metadata,
    pixel_height_to_bandwidth_hz,
    pixel_to_frequency_hz,
    slice_frequency_window,
)


@dataclass
class LegacyAnalysisResult:
    mode: str
    name: str
    path: str | None
    preview_lines: list[str]
    detected_sample_rate_hz: float | None
    sample_rate_hz: float
    input_values: dict[str, Any]
    display_values: dict[str, str]
    estimates: dict[str, float]
    quality: dict[str, float]
    errors: dict[str, float | None]
    signal_samples: Any
    constellation_samples: Any
    images: dict[str, str | None]


@contextlib.contextmanager
def project_cwd():
    previous = Path.cwd()
    os.chdir(PROJECT_ROOT)
    try:
        yield
    finally:
        os.chdir(previous)


@lru_cache(maxsize=1)
def runtime_modules():
    with project_cwd():
        import cv2  # type: ignore
        import numpy as np  # type: ignore
        from PIL import Image  # type: ignore

        from EVM_Ana import signal_ideal
        from image_processing_41 import process_image
        from parameter_tune import flag1_tune, flag2_tune, flag3_tune
        from signal_photo_save import signal_create, signal_read
        from state_process import load_excel_signal
        from unet import Unet

    return {
        "cv2": cv2,
        "np": np,
        "Image": Image,
        "signal_ideal": signal_ideal,
        "process_image": process_image,
        "flag1_tune": flag1_tune,
        "flag2_tune": flag2_tune,
        "flag3_tune": flag3_tune,
        "signal_create": signal_create,
        "signal_read": signal_read,
        "load_excel_signal": load_excel_signal,
        "Unet": Unet,
    }


_UNET_INSTANCE: Any = None


def safe_detect(unet_obj, image):
    result = unet_obj.detect_image(image)
    if isinstance(result, (tuple, list)):
        if len(result) == 3:
            return result[0], result[1], result[2]
        if len(result) == 2:
            return result[0], result[1], None
    return result, None, None


def get_unet():
    global _UNET_INSTANCE
    if _UNET_INSTANCE is None:
        with project_cwd():
            _UNET_INSTANCE = runtime_modules()["Unet"]()
    return _UNET_INSTANCE


def get_model_full_image_path(run_manifest: dict[str, Any]) -> Path:
    candidate = run_manifest.get("model_full_image_path") or run_manifest.get("recognition_image_path")
    if not candidate:
        raise RuntimeError("Current STFT run manifest does not contain a model image path")
    return Path(candidate)


def get_display_source_path(run_manifest: dict[str, Any], fallback: Path) -> Path:
    local_path = run_manifest.get("model_local_image_path")
    if local_path:
        candidate = Path(local_path)
        if candidate.exists():
            return candidate
    return fallback


def derive_psd_band_estimate(source_path: Path) -> dict[str, float]:
    modules = runtime_modules()
    cv2 = modules["cv2"]

    stft_metadata = load_stft_metadata(source_path)
    if not stft_metadata:
        raise RuntimeError(f"STFT metadata is missing for {source_path}")

    image = cv2.imread(str(source_path))
    if image is None:
        raise FileNotFoundError(f"Unable to load STFT image: {source_path}")

    image_height = int(image.shape[0])
    freq_min_hz = float(stft_metadata.get("freq_min_hz", 0.0))
    freq_max_hz = float(stft_metadata.get("freq_max_hz", 0.0))
    psd_band = stft_metadata.get("psd_band") or {}

    band_min_hz = psd_band.get("f_min_3db_hz", psd_band.get("f_min_hz"))
    band_max_hz = psd_band.get("f_max_3db_hz", psd_band.get("f_max_hz"))
    if band_min_hz is None or band_max_hz is None:
        raise RuntimeError("PSD band metadata is incomplete")

    band_min_hz = float(band_min_hz)
    band_max_hz = float(band_max_hz)
    if band_max_hz < band_min_hz:
        band_min_hz, band_max_hz = band_max_hz, band_min_hz

    top_row = frequency_to_pixel_y(band_max_hz, freq_min_hz, freq_max_hz, image_height)
    bottom_row = frequency_to_pixel_y(band_min_hz, freq_min_hz, freq_max_hz, image_height)
    row_start = max(0, min(top_row, bottom_row))
    row_end = min(image_height, max(top_row, bottom_row))
    rect_height = max(1, row_end - row_start)
    center_row = 0.5 * (row_start + row_end)

    return {
        "center_frequency_hz": 0.5 * (band_min_hz + band_max_hz),
        "bandwidth_hz": max(0.0, band_max_hz - band_min_hz),
        "band_min_hz": band_min_hz,
        "band_max_hz": band_max_hz,
        "center_pixel_y": float(center_row),
        "rect_height_px": float(rect_height),
        "image_height_px": float(image_height),
    }


def analyze_file(
    *,
    file_path: Path,
    fs_hz: float | None,
    fc_hz: float,
    rs_hz: float,
    snr_db: float,
    modulation: int,
) -> LegacyAnalysisResult:
    preview_lines = preview_file(file_path)
    rec_wave, effective_fs, detected_fs = load_external_signal(file_path, fs_hz)

    modules = runtime_modules()
    np = modules["np"]
    signal_read = modules["signal_read"]
    signal_ideal = modules["signal_ideal"]
    flag1_tune = modules["flag1_tune"]
    flag2_tune = modules["flag2_tune"]
    flag3_tune = modules["flag3_tune"]

    with project_cwd():
        keep_negative_frequencies = (
            file_path.suffix.lower() == ".txt"
            and np.iscomplexobj(rec_wave)
            and float(np.max(np.abs(np.imag(np.asarray(rec_wave))))) > 1e-12
        )
        Fs, rec_wave, magnitude_estimate, snr_estimate, rs_estimate = signal_read(
            rec_wave,
            effective_fs,
            keep_negative_frequencies=keep_negative_frequencies,
        )

        rs_process = rs_estimate / 1e9
        run_manifest = load_current_stft_run()
        if not run_manifest:
            raise RuntimeError("Current STFT run manifest was not generated")
        image_path_stft = get_model_full_image_path(run_manifest)
        image_path_annotated = Path(run_manifest["annotated_image_path"])
        image_path_display = Path(run_manifest["display_image_path"])
        image_path_display_source = get_display_source_path(run_manifest, image_path_stft)

        psd_estimate = derive_psd_band_estimate(image_path_stft)
        center_frequency_estimate = psd_estimate["center_frequency_hz"]
        bandwidth_estimate = psd_estimate["bandwidth_hz"]
        center_frequency_output = psd_estimate["center_pixel_y"]
        rect_height = int(round(psd_estimate["rect_height_px"]))

        detect_band_region(image_path_stft, run_manifest["mask_image_path"])
        bandwidth_true = (1 + 0.35 / 2) * rs_hz

        evm_percentage, evm_db, papr, signal_star = signal_ideal(
            Fs, center_frequency_estimate, rs_estimate, rec_wave, modulation, snr_estimate
        )

        annotate_stft_image(
            source_path=image_path_stft,
            output_path=image_path_annotated,
            center_frequency_output=center_frequency_output,
            rect_height=rect_height,
        )
        build_stft_display_image(
            source_path=image_path_display_source,
            output_path=image_path_display,
            samples=rec_wave,
            fs_hz=Fs,
        )

    display_values = {
        "center_frequency": scientific_text(center_frequency_estimate),
        "symbol_rate": scientific_text(rs_estimate),
        "bandwidth": scientific_text(bandwidth_estimate),
        "snr": f"{snr_estimate:.2f}",
        "magnitude": f"{magnitude_estimate:.2f}",
        "evm": f"{evm_percentage:.2f}",
        "papr": f"{papr:.2f}",
    }

    estimates = {
        "center_frequency_hz": float(center_frequency_estimate),
        "bandwidth_hz": float(bandwidth_estimate),
        "rs_hz": float(rs_estimate),
        "snr_db": float(snr_estimate),
        "evm_percent": float(evm_percentage),
        "evm_db": float(evm_db),
        "papr": float(papr),
        "magnitude": float(magnitude_estimate),
        "bandwidth_true_hz": float(bandwidth_true),
    }

    quality = calculate_quality_scores(
        center_frequency_estimate=center_frequency_estimate,
        fc_input=fc_hz,
        rs_input=rs_hz,
        bandwidth_estimate=bandwidth_estimate,
        bandwidth_true=bandwidth_true,
        snr_estimate=snr_estimate,
        snr_input=snr_db,
        evm_percent=evm_percentage,
        modulation=modulation,
    )

    errors = calculate_error_metrics(
        center_frequency_estimate=center_frequency_estimate,
        fc_input=fc_hz,
        rs_input=rs_hz,
        bandwidth_estimate=bandwidth_estimate,
        bandwidth_true=bandwidth_true,
        snr_estimate=snr_estimate,
        snr_input=snr_db,
    )

    return LegacyAnalysisResult(
        mode="file",
        name=file_path.name,
        path=str(file_path),
        preview_lines=preview_lines,
        detected_sample_rate_hz=detected_fs,
        sample_rate_hz=float(Fs),
        input_values={
            "fs_hz": float(fs_hz or effective_fs),
            "fc_hz": float(fc_hz),
            "rs_hz": float(rs_hz),
            "snr_db": float(snr_db),
            "modulation": int(modulation),
            "modulation_name": modulation_name(modulation),
        },
        display_values=display_values,
        estimates=estimates,
        quality=quality,
        errors=errors,
        signal_samples=rec_wave,
        constellation_samples=signal_star,
        images={
            "stft_source": path_to_data_url(Path(run_manifest["model_full_image_path"])),
            "stft_annotated": path_to_data_url(image_path_annotated),
            "stft_display": path_to_data_url(image_path_display),
            "mask": path_to_data_url(Path(run_manifest["mask_image_path"])),
            "grad_cam": path_to_data_url(SIGNAL_ANA_DIR / "bandwidth_Grad_Cam.jpg"),
        },
    )


def analyze_generated(
    *,
    fs_hz: float,
    fc_hz: float,
    rs_hz: float,
    snr_db: float,
    modulation: int,
) -> LegacyAnalysisResult:
    modules = runtime_modules()
    signal_create = modules["signal_create"]
    signal_ideal = modules["signal_ideal"]
    flag1_tune = modules["flag1_tune"]
    flag2_tune = modules["flag2_tune"]
    flag3_tune = modules["flag3_tune"]

    with project_cwd():
        Fs, rec_wave, magnitude_estimate, snr_estimate, rs_estimate, min_val, max_val = signal_create(
            fs_hz, fc_hz, rs_hz, snr_db, modulation
        )

        rs_process = rs_estimate / 1e9
        run_manifest = load_current_stft_run()
        if not run_manifest:
            raise RuntimeError("Current STFT run manifest was not generated")
        image_path_stft = get_model_full_image_path(run_manifest)
        image_path_annotated = Path(run_manifest["annotated_image_path"])
        image_path_display = Path(run_manifest["display_image_path"])
        image_path_display_source = get_display_source_path(run_manifest, image_path_stft)

        psd_estimate = derive_psd_band_estimate(image_path_stft)
        center_frequency_estimate = psd_estimate["center_frequency_hz"]
        bandwidth_estimate = psd_estimate["bandwidth_hz"]
        center_frequency_output = psd_estimate["center_pixel_y"]
        rect_height = int(round(psd_estimate["rect_height_px"]))

        detect_band_region(image_path_stft, run_manifest["mask_image_path"])
        bandwidth_true = (1 + 0.35) * rs_hz

        evm_percentage, evm_db, papr, signal_star = signal_ideal(
            Fs, center_frequency_estimate, rs_estimate, rec_wave, modulation, snr_estimate
        )

        annotate_stft_image(
            source_path=image_path_stft,
            output_path=image_path_annotated,
            center_frequency_output=center_frequency_output,
            rect_height=rect_height,
        )
        build_stft_display_image(
            source_path=image_path_display_source,
            output_path=image_path_display,
            samples=rec_wave,
            fs_hz=Fs,
        )

    display_values = {
        "center_frequency": scientific_text(center_frequency_estimate),
        "symbol_rate": scientific_text(rs_estimate),
        "bandwidth": scientific_text(bandwidth_estimate),
        "snr": f"{snr_estimate:.2f}",
        "magnitude": f"{magnitude_estimate:.2f}",
        "evm": f"{evm_percentage:.2f}",
        "papr": f"{papr:.2f}",
    }

    estimates = {
        "center_frequency_hz": float(center_frequency_estimate),
        "bandwidth_hz": float(bandwidth_estimate),
        "rs_hz": float(rs_estimate),
        "snr_db": float(snr_estimate),
        "evm_percent": float(evm_percentage),
        "evm_db": float(evm_db),
        "papr": float(papr),
        "magnitude": float(magnitude_estimate),
        "bandwidth_true_hz": float(bandwidth_true),
        "stft_min": float(min_val),
        "stft_max": float(max_val),
    }

    quality = calculate_quality_scores(
        center_frequency_estimate=center_frequency_estimate,
        fc_input=fc_hz,
        rs_input=rs_hz,
        bandwidth_estimate=bandwidth_estimate,
        bandwidth_true=bandwidth_true,
        snr_estimate=snr_estimate,
        snr_input=snr_db,
        evm_percent=evm_percentage,
        modulation=modulation,
    )

    errors = calculate_error_metrics(
        center_frequency_estimate=center_frequency_estimate,
        fc_input=fc_hz,
        rs_input=rs_hz,
        bandwidth_estimate=bandwidth_estimate,
        bandwidth_true=bandwidth_true,
        snr_estimate=snr_estimate,
        snr_input=snr_db,
    )

    preview_lines = [
        f"Fs={fs_hz}",
        f"Fc={fc_hz}",
        f"Rs={rs_hz}",
        f"SNR={snr_db}",
        f"Modulation={modulation_name(modulation)}",
    ]

    return LegacyAnalysisResult(
        mode="generated",
        name=f"generated_{modulation_name(modulation).lower()}",
        path=None,
        preview_lines=preview_lines,
        detected_sample_rate_hz=None,
        sample_rate_hz=float(Fs),
        input_values={
            "fs_hz": float(fs_hz),
            "fc_hz": float(fc_hz),
            "rs_hz": float(rs_hz),
            "snr_db": float(snr_db),
            "modulation": int(modulation),
            "modulation_name": modulation_name(modulation),
        },
        display_values=display_values,
        estimates=estimates,
        quality=quality,
        errors=errors,
        signal_samples=rec_wave,
        constellation_samples=signal_star,
        images={
            "stft_source": path_to_data_url(Path(run_manifest["model_full_image_path"])),
            "stft_annotated": path_to_data_url(image_path_annotated),
            "stft_display": path_to_data_url(image_path_display),
            "mask": path_to_data_url(Path(run_manifest["mask_image_path"])),
            "grad_cam": path_to_data_url(SIGNAL_ANA_DIR / "bandwidth_Grad_Cam.jpg"),
        },
    )


def preview_file(file_path: Path) -> list[str]:
    suffix = file_path.suffix.lower()
    if suffix in {".xlsx", ".xls"}:
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception:
            return ["Excel preview unavailable: openpyxl is not installed"]

        workbook = load_workbook(file_path, data_only=True, read_only=True)
        sheet = workbook.active
        preview_lines: list[str] = []
        for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_index >= 5:
                break
            preview_lines.append("\t".join("" if cell is None else str(cell) for cell in row))
        return preview_lines

    preview_lines = []
    with file_path.open("r", encoding="utf-8", errors="replace") as file:
        for line_index, line in enumerate(file):
            if line_index >= 5:
                break
            preview_lines.append(line.rstrip("\r\n"))
    return preview_lines


def load_external_signal(file_path: Path, fs_hz: float | None):
    modules = runtime_modules()
    np = modules["np"]
    load_excel_signal = modules["load_excel_signal"]
    suffix = file_path.suffix.lower()

    with project_cwd():
        if suffix in {".xlsx", ".xls", ".csv"}:
            rec_wave, file_sample_rate = load_excel_signal(str(file_path), normalize=True, return_fs=True)
            effective_fs = file_sample_rate or fs_hz
            if effective_fs is None:
                raise ValueError("Sampling rate could not be detected from the file and Fs was not provided")
            return rec_wave, float(effective_fs), float(file_sample_rate) if file_sample_rate is not None else None

        if suffix == ".wfm":
            raise ValueError(
                "Raw .wfm binary files are not supported yet. Please export the waveform as CSV or TXT first."
            )

        rec_wave = parse_text_signal(file_path, np)
        if fs_hz is None:
            raise ValueError("Text signal file requires a manual Fs input")
        return rec_wave, float(fs_hz), None


FLOAT_PATTERN = re.compile(r"[-+]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][-+]?\d+)?")


def parse_text_signal(file_path: Path, np):
    """Accept complex-per-line TXT plus simple numeric column formats."""
    complex_values: list[complex] = []
    numeric_rows: list[list[float]] = []
    invalid_examples: list[str] = []

    lines = file_path.read_text(encoding="utf-8", errors="replace").splitlines()
    for raw_line in lines:
        line = raw_line.strip().lstrip("\ufeff")
        if not line or line.startswith("#") or line.startswith("//"):
            continue

        candidate = line.replace("I", "j").replace("i", "j")
        compact_candidate = candidate.replace(" ", "").replace("\t", "")
        compact_candidate = compact_candidate.rstrip(",;")
        tokens = FLOAT_PATTERN.findall(line)

        if "j" in compact_candidate.lower():
            try:
                complex_values.append(complex(compact_candidate))
                continue
            except ValueError:
                pass

        if len(tokens) >= 2:
            numeric_rows.append([float(token) for token in tokens])
        elif len(tokens) == 1:
            complex_values.append(complex(float(tokens[0]), 0.0))
        elif len(invalid_examples) < 3:
            invalid_examples.append(line[:120])

    if complex_values and not numeric_rows:
        return np.asarray(complex_values, dtype=np.complex128)

    if numeric_rows:
        usable_rows = [row for row in numeric_rows if row]
        if not usable_rows:
            raise ValueError("No numeric samples were found")

        column_count = max(len(row) for row in usable_rows)
        if column_count >= 2:
            two_column_rows = [row for row in usable_rows if len(row) >= 2]
            real_signal = np.asarray([row[0] for row in two_column_rows], dtype=float)
            imag_signal = np.asarray([row[1] for row in two_column_rows], dtype=float)
            return real_signal + 1j * imag_signal

        real_signal = np.asarray([row[0] for row in usable_rows], dtype=float)
        try:
            from scipy.signal import hilbert  # type: ignore
        except Exception as exc:
            raise ValueError("Single-column real TXT files require scipy.signal.hilbert") from exc
        return hilbert(real_signal)

    details = f" Invalid rows: {invalid_examples}" if invalid_examples else ""
    raise ValueError(
        "Unable to parse text signal file. Supported TXT formats: one complex sample per line, "
        "two numeric I/Q columns, or one numeric real-signal column."
        + details
    )


def detect_sample_rate(file_path: Path) -> float | None:
    modules = runtime_modules()
    load_excel_signal = modules["load_excel_signal"]

    if file_path.suffix.lower() not in {".xlsx", ".xls", ".csv"}:
        return None

    with project_cwd():
        _, file_sample_rate = load_excel_signal(str(file_path), normalize=True, return_fs=True)
    return float(file_sample_rate) if file_sample_rate is not None else None


def detect_band_region(image_path: str | Path, mask_output_path: str | Path | None = None) -> int:
    modules = runtime_modules()
    cv2 = modules["cv2"]
    Image = modules["Image"]

    with project_cwd():
        image = cv2.imread(str(image_path))
        if image is None:
            raise FileNotFoundError(f"Unable to load STFT image: {image_path}")
        image_height = image.shape[0]
        unet = get_unet()
        pil_image = Image.open(str(image_path))
        rendered_image, rect_height, _ = safe_detect(unet, pil_image)
        if rect_height is None:
            rect_height = 0
        try:
            dynamic_mask_path = Path(mask_output_path) if mask_output_path else Path(image_path).with_name(
                f"{Path(image_path).stem}_unet_mask.jpg"
            )
            rendered_image.save(dynamic_mask_path)
        except Exception:
            pass
        return int(rect_height if rect_height else image_height * 0.2)


def build_bandwidth_estimate(
    *,
    rect_height: int,
    image_path: str | Path,
    rs_process: float,
    snr_estimate: float,
    fc_process: float,
    flag1_tune,
    flag2_tune,
    flag3_tune,
) -> float:
    modules = runtime_modules()
    cv2 = modules["cv2"]
    image = cv2.imread(str(image_path))
    if image is None:
        raise FileNotFoundError(f"Unable to load STFT image: {image_path}")

    image_height = image.shape[0]
    stft_metadata = load_stft_metadata(image_path)
    if stft_metadata:
        bandwidth_estimate_1 = pixel_height_to_bandwidth_hz(
            rect_height,
            stft_metadata,
            image_height_px=image_height,
        ) / 1e9
        bandwidth_estimate_2 = bandwidth_estimate_1
    else:
        bandwidth_estimate_1 = rect_height / image_height * 10
        bandwidth_estimate_2 = rect_height / image_height * 10
    band_flag = compute_band_flag(rs_process)

    return (
        (bandwidth_estimate_1 + bandwidth_estimate_2)
        / 2
        * flag3_tune(rs_process)
        / band_flag
        * 1e9
        / flag1_tune(rs_process, snr_estimate, fc_process)
        / flag2_tune(rs_process, snr_estimate, fc_process)
    )


def compute_band_flag(rs_process: float) -> float:
    rounded = round(rs_process, 2)
    if rounded < 0.745:
        return 0.753124
    if 0.745 <= rounded < 0.99:
        return 0.755928
    if 0.99 <= rounded < 1.20:
        return 0.733674
    if 1.20 <= rounded < 1.35:
        return 0.727199
    if 1.35 <= rounded < 1.55:
        return 0.6722442
    if 1.55 <= rounded < 1.85:
        return 0.619473
    if 1.85 <= rounded < 2.05:
        return 0.610236
    if 2.05 <= rounded < 2.65:
        return 0.695199
    return 1.0


def annotate_stft_image(
    *,
    source_path: Path,
    output_path: Path,
    center_frequency_output: float,
    rect_height: int,
):
    modules = runtime_modules()
    cv2 = modules["cv2"]

    image = cv2.imread(str(source_path))
    if image is None:
        return

    height, width, _ = image.shape
    output_image = image.copy()
    point_top = (20, int(center_frequency_output - 0.5 * rect_height))
    point_bottom = (width - 20, int(center_frequency_output + 0.5 * rect_height))
    cv2.rectangle(output_image, point_top, point_bottom, (0, 255, 0), 8)
    cv2.circle(output_image, (int(width * 0.5), int(center_frequency_output)), 5, (255, 0, 0), -1)
    cv2.imwrite(str(output_path), output_image)


def build_stft_display_image(*, source_path: Path, output_path: Path, samples=None, fs_hz: float | None = None):
    if samples is not None and fs_hz:
        if build_stft_plot_image(samples=samples, fs_hz=fs_hz, output_path=output_path, source_path=source_path):
            return

    modules = runtime_modules()
    Image = modules["Image"]

    try:
        with Image.open(source_path) as image:
            resampling = getattr(getattr(Image, "Resampling", Image), "BICUBIC")
            rotated = image.rotate(-90, expand=True).resize(image.size, resampling)
            rotated.save(output_path, quality=92)
    except Exception:
        return


def build_stft_plot_image(*, samples, fs_hz: float, output_path: Path, source_path: Path | None = None) -> bool:
    try:
        import matplotlib

        matplotlib.use("Agg", force=True)
        import matplotlib.pyplot as plt
        import numpy as np  # type: ignore
        from matplotlib.font_manager import FontProperties
        from matplotlib.ticker import ScalarFormatter
        from mpl_toolkits.axes_grid1.inset_locator import inset_axes

        signal = np.asarray(samples, dtype=np.complex128)
        if signal.size < 512:
            return False

        max_samples = 4_000_000
        if signal.size > max_samples:
            signal = signal[:max_samples]

        fig, ax = plt.subplots(figsize=(16, 3.7), dpi=220)
        fig.patch.set_facecolor("black")
        ax.set_facecolor("black")

        stft_metadata = load_stft_metadata(source_path) if source_path else None
        keep_negative_frequencies = bool(stft_metadata.get("keep_negative_frequencies")) if stft_metadata else False
        display_min_db = stft_metadata.get("display_min_db") if stft_metadata else None
        display_max_db = stft_metadata.get("display_max_db") if stft_metadata else None

        freqs, times, matrix, _ = compute_spectrogram_matrix(
            signal,
            fs_hz,
            keep_negative_frequencies=keep_negative_frequencies,
        )
        if stft_metadata and stft_metadata.get("role") == "local":
            freqs, matrix = slice_frequency_window(
                freqs,
                matrix,
                float(stft_metadata.get("freq_min_hz", float(freqs.min()))),
                float(stft_metadata.get("freq_max_hz", float(freqs.max()))),
            )
        if freqs.size == 0 or matrix.size == 0:
            return False

        image = ax.imshow(
            matrix.T,
            aspect="auto",
            origin="lower",
            cmap="jet",
            extent=[float(freqs.min()), float(freqs.max()), float(times.min()), float(times.max())],
            vmin=display_min_db,
            vmax=display_max_db,
        )

        font = FontProperties(family="Segoe UI", size=12)
        ax.set_xlabel("Frequency (Hz)", color="white", fontproperties=font, labelpad=-1)
        ax.set_ylabel("Time (s)", color="white", fontproperties=font)
        ax.xaxis.set_major_formatter(ScalarFormatter(useMathText=True))
        ax.xaxis.get_major_formatter().set_powerlimits((0, 1))
        ax.yaxis.set_major_formatter(ScalarFormatter(useMathText=True))
        ax.yaxis.get_major_formatter().set_powerlimits((0, 1))
        ax.tick_params(colors="white")
        for label in ax.get_xticklabels() + ax.get_yticklabels():
            label.set_fontproperties(font)

        cax = inset_axes(ax, width="100%", height="2.5%", loc="lower center", borderpad=-3.5)
        colorbar = fig.colorbar(image, cax=cax, orientation="horizontal")
        colorbar.ax.tick_params(labelsize=10, colors="white")

        fig.subplots_adjust(left=0.04, right=0.995, top=0.94, bottom=0.2)
        fig.savefig(output_path, facecolor="black")
        plt.close(fig)
        return True
    except Exception:
        return False


def calculate_quality_scores(
    *,
    center_frequency_estimate: float,
    fc_input: float,
    rs_input: float,
    bandwidth_estimate: float,
    bandwidth_true: float,
    snr_estimate: float,
    snr_input: float,
    evm_percent: float,
    modulation: int,
) -> dict[str, float]:
    center_frequency_quality = 25 - ((abs(center_frequency_estimate - fc_input) / rs_input) * 20 - 1)
    bandwidth_quality = 25 - ((abs(bandwidth_estimate - bandwidth_true) / rs_input) * 10 - 1)
    snr_quality = 25 - (abs(snr_estimate - snr_input) * 2.5)

    if modulation == 1:
        evm_quality = 25 - ((evm_percent / 17.5) * 10 - 1) * 5
    elif modulation == 2:
        evm_quality = 25 - ((evm_percent / 12) * 10 - 1) * 5
    elif modulation == 3:
        evm_quality = 25 - ((evm_percent / 12.5) * 10 - 1) * 5
    elif modulation == 4:
        evm_quality = 25 - ((evm_percent / 8) * 10 - 1) * 5
    else:
        evm_quality = 15

    return {
        "snr": clamp(snr_quality, 0, 25),
        "center_frequency": clamp(center_frequency_quality, 0, 25),
        "bandwidth": clamp(bandwidth_quality, 0, 25),
        "evm": clamp(evm_quality, 0, 25),
    }


def calculate_error_metrics(
    *,
    center_frequency_estimate: float,
    fc_input: float,
    rs_input: float,
    bandwidth_estimate: float,
    bandwidth_true: float,
    snr_estimate: float,
    snr_input: float,
) -> dict[str, float | None]:
    return {
        "center_frequency_percent": abs((center_frequency_estimate - fc_input) / rs_input * 100) if rs_input else None,
        "bandwidth_percent": abs((bandwidth_estimate - bandwidth_true) / rs_input * 100) if rs_input else None,
        "snr_db": abs(snr_estimate - snr_input),
    }


def path_to_data_url(path: Path) -> str | None:
    if not path.exists() or not path.is_file():
        return None
    mime_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
    encoded = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:{mime_type};base64,{encoded}"


def scientific_text(value: float) -> str:
    return f"{value:.2e}"


def modulation_name(modulation: int) -> str:
    return {
        1: "QPSK",
        2: "8PSK",
        3: "16QAM",
        4: "64QAM",
    }.get(modulation, f"Modulation {modulation}")


def clamp(value: float, lower: float, upper: float) -> float:
    return max(lower, min(upper, value))
