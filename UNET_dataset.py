import os
import cv2
import numpy as np
import json
import base64

import openpyxl
from PIL import Image
from unet import Unet
import os

from signal_photo_save import signal_create_test
from image_processing_41 import process_image_test

def parse_coordinates_from_filename(filename):
    """从文件名中解析出深红色区域的坐标."""
    # 去掉文件扩展名，如 'img_10_20_50_60.jpg' -> 'img_10_20_50_60'
    name = os.path.splitext(filename)[0]

    # 提取坐标（假设格式：img_x1_y1_x2_y2）
    _, x1, y1, x2, y2 = map(int, name.split('_'))
    return x1, y1, x2, y2

def encode_image_to_base64(image_path):
    """将图像读取并编码为 Base64 字符串."""
    with open(image_path, "rb") as image_file:
        image_data = base64.b64encode(image_file.read()).decode('utf-8')
    return image_data
def generate_json_from_coordinates(image_dir, json_dir, x1, x2, y1, y2):
    """根据文件名中的坐标生成 JSON 标注文件."""
    # 读取图像尺寸
    image = cv2.imread(image_dir)
    height, width = image.shape[:2]
    #
    # # 从文件名解析坐标
    # x1, y1, x2, y2 = parse_coordinates_from_filename(filename)

    # 编码图像为 Base64
    image_data = encode_image_to_base64(image_dir)

    # 构建 JSON 数据结构
    annotation = {
        "version": "5.5.0",
        "flags": {},
        "shapes": [
            {
                "label": "red_region",
                "points": [[x1, y1], [x2, y1], [x2, y2], [x1, y2]],
                "group_id": None,
                "description": "",
                "shape_type": "polygon",
                "flags": {},
                "mask": None
            }
        ],
        "imagePath": os.path.basename(image_dir),
        "imageData": image_data,
        "imageHeight": height,
        "imageWidth": width
    }
    # 将 JSON 数据写入文件
    with open(json_dir, 'w') as f:
        json.dump(annotation, f, indent=4)
    filename = os.path.basename(image_dir)
    print(f'Saved JSON for {filename}')

# 批量生成掩膜图
# 循环遍历
for i in np.arange(0.5, 2.5, 0.25):
    for j in range(1, 11):
        for k in range(1, 201, 10):
            for Modulation in range(1, 2):
                # 调用函数，计算相关参数
                Fs = 10 * 1e9
                Fs_index = Fs / 1e9
                Rs_index = i
                Fc = 3.0 + j * 0.05 - 0.05
                SNR = -10 + (k - 1) * 0.1
                Idenx = 200 * ((i - 0.5) / 0.25) + (j - 1) * 20 + (k - 1) / 10 + Modulation
                Fc_index = round(Fc, 2)
                SNR_index = round(SNR, 1)
                image_dir = f'./dataset_YOLO/image_dataset.jpg'  # 替换为你的图像文件夹路径

                [Fs, rec_wave, SNR_GUJI, RS_GUJI] = signal_create_test(Fs, Fc_index * 1e9, Rs_index * 1e9, SNR_index, Modulation, image_dir)
                if (round(Rs_index, 2) < 0.98):
                    Band_flag = 1.07
                elif (0.99 <= round(Rs_index, 2) < 1.20):
                    Band_flag = 1.04
                elif (1.20 <= round(Rs_index, 2) < 1.35):
                    Band_flag = 1.035
                elif (1.35 <= round(Rs_index, 2) < 1.55):
                    Band_flag = 0.958
                elif (1.55 <= round(Rs_index, 2) < 1.85):
                    Band_flag = 0.882
                elif (1.85 <= round(Rs_index, 2) < 2.05):
                    Band_flag = 0.867
                elif (2.05 <= round(Rs_index, 2) < 2.65):
                    Band_flag = 0.992
                else:
                    Band_flag = 1
                [height, center_frequency_output] = process_image_test(rec_wave, Fs, image_dir)

                unet = Unet()
                image = Image.open(image_dir)
                r_image, rect_height = unet.detect_image(image)

                file_path = f'./Expected_parameter.xlsx'

                wb1 = openpyxl.load_workbook(file_path)
                sheet1 = wb1["Sheet1"]

                # sheet = wb.active
                # Convert row_number to 0-based index for openpyxl
                row_index = int(Idenx)
                # Write data to specified row and columns A, B, C
                sheet1.cell(row=row_index + 1, column=3).value = Fc_index
                sheet1.cell(row=row_index + 1, column=4).value = Rs_index
                sheet1.cell(row=row_index + 1, column=6).value = SNR_index
                wb1.save(file_path)

                wb2 = openpyxl.load_workbook(file_path)
                sheet2 = wb2["Sheet2"]
                # sheet2.cell(row=row_index + 1, column=1).value = height
                sheet2.cell(row=row_index + 1, column=2).value = rect_height / Band_flag
                sheet2.cell(row=row_index + 1, column=3).value = center_frequency_output
                sheet2.cell(row=row_index + 1, column=6).value = SNR_GUJI
                sheet2.cell(row=row_index + 1, column=7).value = (RS_GUJI / 1e9)
                wb2.save(file_path)

                print(f"Data successfully written to row {Idenx}")
                # image_dir = f'./dataset_YOLO/images/image_{image_id}.jpg'  # 替换为你的图像文件夹路径
                # json_dir = f'./dataset_YOLO/mask/image_{image_id}.json' # 替换为掩膜保存路径
                # generate_json_from_coordinates(image_dir, json_dir, x1, x2, y1, y2)
